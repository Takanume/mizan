"""Génère un Excel de corrections de date pré-rempli avec les factures AN.

Le cabinet ouvre ensuite ce fichier et remplit la colonne "Date Facture réelle"
en s'appuyant sur les PDF justificatifs. Mizan reprend ces dates au prochain
`mizan run` pour recalculer les retards correctement.

Usage :
    python3 scripts/generer_corrections_dates.py

Produit :
    out/Corrections Dates UEMA 1T26.xlsx
"""

from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from parser import parser_gl_liste  # noqa: E402
from lettrage import construire_lettrages  # noqa: E402
from compute import calculer_toutes_lignes, charger_base_fournisseurs  # noqa: E402
from models import StatutFacture  # noqa: E402


ROOT = Path(__file__).resolve().parents[1]
GL_UEMA  = ROOT / "samples" / "input" / "UEMA - GL FRS 2026.xlsx"
TEMPLATE = ROOT / "samples" / "input" / "Modèle Suivi Global.xlsx"
REF_DGI  = ROOT / "samples" / "input" / "Référentiel DGI - Cabinet.xlsx"
SORTIE   = ROOT / "out" / "Corrections Dates UEMA 1T26.xlsx"

# Couleurs (charte Mizan)
NAVY = "FF2B4C6F"
TEAL = "FF4ECDC4"
PAPER = "FFF8FAFC"
YELLOW = "FFFFF3CD"
LINE_COLOR = "FFD8E0EA"

HEADERS = [
    ("N° Fournisseur",         14),  # A
    ("Nom Fournisseur",        30),  # B
    ("N° Facture",             18),  # C
    ("Date Facture réelle",    20),  # D (à remplir par le cabinet)
    ("Date Livraison",         20),  # E (optionnel, défaut = Date Facture)
    ("Observations",           30),  # F
    ("Délai spécifique (j)",   18),  # G (optionnel — surcharge le délai fournisseur)
]

# Colonnes en lecture seule (à pré-remplir par Mizan pour aider le cabinet)
INFOS_LECTURES = [
    ("Date Sage (à remplacer)", 22),  # G
    ("Délai (j)",                12),  # H
    ("Montant TTC",              16),  # I
    ("Statut actuel",            28),  # J
]


def _detecter_millesime_an(n_facture: str | None, annee_courante: int) -> str:
    """Si le n° facture commence par 'NN/' ou 'NN-', détecte l'année probable."""
    if not n_facture:
        return ""
    m = re.match(r"^(\d{2})[^\d]", n_facture.strip())
    if not m:
        return ""
    millesime = int(m.group(1))
    if millesime < 80:
        annee = 2000 + millesime
    else:
        annee = 1900 + millesime
    if annee != annee_courante:
        return f"Millésime {annee} probable"
    return ""


def generer():
    print("→ Chargement de la base...")
    base = charger_base_fournisseurs(REF_DGI)
    print(f"  {len(base)} fournisseurs")

    print("→ Parsing du GL...")
    ecritures = parser_gl_liste(GL_UEMA)
    print(f"  {len(ecritures)} écritures")

    print("→ Construction des lettrages...")
    lettrages, _, _ = construire_lettrages(ecritures, base_fournisseurs=base)
    print(f"  {len(lettrages)} lettrages")

    print("→ Calcul des lignes de suivi...")
    lignes = calculer_toutes_lignes(lettrages, debut_periode=date(2026, 1, 1))

    # Filtrer les factures à corriger = celles flaggées par le patch D-006
    # (factures reportées AN ou avec millésime ancien dans le n°)
    a_corriger = [
        l for l in lignes
        if l.observations and l.observations.startswith("⚠️")
    ]
    print(f"  {len(a_corriger)} factures candidates à correction")

    # Trier par fournisseur puis date d'écriture
    a_corriger.sort(key=lambda l: (l.fournisseur, l.date_facture, l.n_facture or ""))

    # ─── Création de l'Excel ─────────────────────────────────────────────
    SORTIE.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corrections Dates"

    nb_cols = len(HEADERS) + len(INFOS_LECTURES)

    # Titre
    ws.cell(1, 1, "CORRECTIONS DE DATES — UEMA 1T26")
    ws.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    ws.row_dimensions[1].height = 28

    # Notice
    ws.cell(2, 1, (
        "Remplir la colonne 'Date Facture réelle' (col D) depuis les PDF justificatifs. "
        "Optionnel : 'Délai spécifique' (col G) pour surcharger le délai du fournisseur "
        "sur cette facture. Mizan recalculera échéance et retards. "
        "Les colonnes grisées sont en lecture seule (référence)."
    ))
    ws.cell(2, 1).font = Font(italic=True, color="FF64748B")
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nb_cols)
    ws.row_dimensions[2].height = 36

    # En-tête colonnes (ligne 3)
    header_row = 3
    col_idx = 1
    # Colonnes éditables (à remplir par le cabinet)
    for label, width in HEADERS:
        c = ws.cell(header_row, col_idx, label)
        c.font = Font(bold=True, color="FFFFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = width
        col_idx += 1
    # Colonnes lecture seule (référence)
    for label, width in INFOS_LECTURES:
        c = ws.cell(header_row, col_idx, label + " (RO)")
        c.font = Font(bold=True, color="FFFFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="FF64748B")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = width
        col_idx += 1
    ws.row_dimensions[header_row].height = 36

    # Bordures
    thin = Side(border_style="thin", color=LINE_COLOR)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Lignes de données
    row = header_row + 1
    nb_avec_millesime = 0
    for ligne in a_corriger:
        # Colonnes éditables
        ws.cell(row, 1, ligne.code_fournisseur or "")
        ws.cell(row, 2, ligne.fournisseur)
        ws.cell(row, 3, ligne.n_facture or "")
        # Col D (Date Facture réelle) — vide à remplir
        # Col E (Date Livraison) — vide à remplir si différente
        millesime_hint = _detecter_millesime_an(ligne.n_facture, 2026)
        if millesime_hint:
            nb_avec_millesime += 1
        ws.cell(row, 6, millesime_hint or "Date à vérifier")

        # Colonnes lecture seule (décalées d'1 col à cause de la nouvelle col G)
        ws.cell(row, 8, ligne.date_facture)
        ws.cell(row, 9, ligne.delai_convenu_jours)
        ws.cell(row, 10, float(ligne.montant_ttc))
        ws.cell(row, 11, ligne.statut.value)

        # Style
        for col in range(1, nb_cols + 1):
            c = ws.cell(row, col)
            c.border = cell_border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            # Colonnes à remplir : fond jaune pâle pour signaler
            if col in (4, 5, 7):
                c.fill = PatternFill("solid", fgColor=YELLOW)
            # Colonnes lecture seule (à partir de col 8 — décalé d'1 à cause de G) : fond gris clair
            elif col >= 8:
                c.fill = PatternFill("solid", fgColor="FFEEF2F7")
                c.font = Font(color="FF64748B", italic=True)
            else:
                c.fill = PatternFill("solid", fgColor=PAPER if (row - header_row) % 2 == 0 else "FFFFFFFF")

        # Formats
        ws.cell(row, 8).number_format = "dd/mm/yyyy"
        ws.cell(row, 4).number_format = "dd/mm/yyyy"
        ws.cell(row, 5).number_format = "dd/mm/yyyy"
        ws.cell(row, 10).number_format = "#,##0.00"

        row += 1

    # Figer les volets
    ws.freeze_panes = ws.cell(header_row + 1, 4)

    wb.save(SORTIE)

    print(f"\n✓ Fichier généré : {SORTIE}")
    print(f"  Total factures à corriger     : {len(a_corriger)}")
    print(f"  Avec millésime ancien détecté : {nb_avec_millesime}")
    print(f"  À compléter par le cabinet    : col D 'Date Facture réelle'")


if __name__ == "__main__":
    generer()
