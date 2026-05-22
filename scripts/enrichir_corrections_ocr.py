"""Sprint 2b — Enrichissement OCR du fichier de corrections de date.

Scanne un dossier de PDF justificatifs, identifie les factures, extrait la
date et le n° via OCR, et enrichit le fichier `Corrections Dates UEMA.xlsx`
avec les dates trouvées (le cabinet n'a plus qu'à valider).

Usage :
    python3 scripts/enrichir_corrections_ocr.py [dossier_pdf] [fichier_corrections]

Par défaut :
    dossier_pdf  = samples/input/Sondage/
    fichier_corr = out/Corrections Dates UEMA 1T26.xlsx
"""

from __future__ import annotations

import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from ocr import TypeDocument, extraire_dossier  # noqa: E402


ROOT = Path(__file__).resolve().parents[1]
DOSSIER_PDF_DEFAUT  = ROOT / "samples" / "input" / "Sondage"
FICHIER_CORRECTIONS = ROOT / "out" / "Corrections Dates UEMA 1T26.xlsx"

# Mêmes couleurs que le générateur de corrections
NAVY      = "FF2B4C6F"
TEAL_LIGHT = "FFD4EFE7"   # vert clair pour les lignes auto-remplies par OCR
YELLOW    = "FFFFF3CD"
LINE      = "FFD8E0EA"


def enrichir(dossier_pdf: Path, fichier_corrections: Path) -> dict[str, int]:
    """Lance l'OCR sur le dossier PDF et écrit les résultats dans le fichier.

    Pour chaque facture détectée, on ajoute (ou met à jour) une ligne :
      - col D (Date Facture réelle) : pré-remplie depuis OCR
      - col E (Date Livraison)      : laissée vide (= date facture par défaut)
      - col F (Observations)        : "OCR conf=NN%, page X de FICHIER.pdf"
    """
    print(f"→ OCR sur {dossier_pdf}...")
    t0 = time.time()
    resultats = extraire_dossier(dossier_pdf)
    factures_brutes = [r for r in resultats if r.type_document == TypeDocument.FACTURE and r.date_facture]

    # Filtres de plausibilité — l'OCR peut lire un "2026" comme "2017" ou "2003".
    # On rejette les résultats peu fiables pour ne pas contaminer le calcul d'échéance.
    from datetime import date as _date, timedelta
    SEUIL_CONFIANCE_MIN = 0.50
    FENETRE_PASSE_JOURS = 365 * 3   # max 3 ans dans le passé
    aujourd_hui = _date.today()
    date_min = aujourd_hui - timedelta(days=FENETRE_PASSE_JOURS)
    date_max = aujourd_hui + timedelta(days=30)

    factures = []
    rejets_confiance = rejets_date = 0
    for r in factures_brutes:
        if r.confiance < SEUIL_CONFIANCE_MIN:
            rejets_confiance += 1
            continue
        if r.date_facture < date_min or r.date_facture > date_max:
            rejets_date += 1
            continue
        factures.append(r)

    print(f"  {time.time()-t0:.1f}s — {len(factures_brutes)} factures détectées, "
          f"{len(factures)} retenues après filtres "
          f"(rejets : confiance<{SEUIL_CONFIANCE_MIN:.0%}={rejets_confiance}, "
          f"date hors fenêtre={rejets_date})")

    if not factures:
        return {"trouvees": 0, "ajoutees": 0, "mises_a_jour": 0}

    # Charger le fichier de corrections existant
    if not fichier_corrections.exists():
        raise FileNotFoundError(
            f"Fichier corrections introuvable : {fichier_corrections}\n"
            f"Génère-le d'abord avec scripts/generer_corrections_dates.py"
        )
    wb = openpyxl.load_workbook(fichier_corrections)
    ws = wb["Corrections Dates"]

    # Construire un index des lignes existantes par (nom, n°)
    header_row = None
    for i in range(1, min(ws.max_row, 10) + 1):
        if str(ws.cell(i, 1).value or "").lower().startswith("n°"):
            header_row = i
            break
    if header_row is None:
        header_row = 3

    index = {}
    derniere_ligne = header_row
    for i in range(header_row + 1, ws.max_row + 1):
        nom = ws.cell(i, 2).value
        n_f = ws.cell(i, 3).value
        if nom and n_f:
            key = (str(nom).strip().upper(), str(n_f).strip().upper())
            index[key] = i
            derniere_ligne = i

    # Style pour les lignes auto-OCR
    thin = Side(border_style="thin", color=LINE)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    stats = {"trouvees": len(factures), "ajoutees": 0, "mises_a_jour": 0, "deja_renseignees": 0}

    for r in factures:
        # Déduire le nom du fournisseur depuis le nom du PDF
        nom_pdf = Path(r.chemin_pdf).stem.upper()
        # Strip suffixes ".PDF" éventuels
        nom = nom_pdf

        n_fact = r.n_facture
        if not n_fact:
            # Pas de n° trouvé → on ne peut pas créer une ligne ciblée
            # On l'ajoute en bas avec n° = "?"
            n_fact = "(?)"

        key = (nom, n_fact.upper())

        if key in index:
            row = index[key]
            # Si la date est déjà saisie manuellement par le cabinet, on ne touche pas
            existing = ws.cell(row, 4).value
            if existing:
                stats["deja_renseignees"] += 1
                continue
            # Sinon on remplit
            ws.cell(row, 4).value = r.date_facture
            ws.cell(row, 4).number_format = "dd/mm/yyyy"
            ws.cell(row, 6).value = (
                f"OCR conf={r.confiance:.0%}, page {r.page} de {Path(r.chemin_pdf).name}"
            )
            # Surligner la ligne en vert clair
            for col in range(1, 11):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=TEAL_LIGHT)
            stats["mises_a_jour"] += 1
        else:
            # Nouvelle ligne (le fichier de corrections n'avait pas cette facture)
            derniere_ligne += 1
            row = derniere_ligne
            ws.cell(row, 1).value = ""   # code fournisseur inconnu
            ws.cell(row, 2).value = nom
            ws.cell(row, 3).value = n_fact
            ws.cell(row, 4).value = r.date_facture
            ws.cell(row, 4).number_format = "dd/mm/yyyy"
            ws.cell(row, 5).value = None
            ws.cell(row, 5).number_format = "dd/mm/yyyy"
            ws.cell(row, 6).value = (
                f"AJOUTÉ OCR conf={r.confiance:.0%}, page {r.page} de {Path(r.chemin_pdf).name}"
            )
            for col in range(1, 11):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=TEAL_LIGHT)
                ws.cell(row, col).border = cell_border
                ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            stats["ajoutees"] += 1

    wb.save(fichier_corrections)
    return stats


def main():
    dossier = Path(sys.argv[1]) if len(sys.argv) > 1 else DOSSIER_PDF_DEFAUT
    fichier = Path(sys.argv[2]) if len(sys.argv) > 2 else FICHIER_CORRECTIONS

    print(f"Dossier PDF       : {dossier}")
    print(f"Fichier corrections : {fichier}\n")

    stats = enrichir(dossier, fichier)

    print(f"\n✓ Enrichissement terminé")
    print(f"  Factures OCR détectées : {stats['trouvees']}")
    print(f"  Lignes mises à jour    : {stats['mises_a_jour']}  (col D pré-remplie)")
    print(f"  Lignes ajoutées        : {stats['ajoutees']}     (nouvelles entrées)")
    print(f"  Lignes déjà renseignées : {stats['deja_renseignees']} (saisie manuelle conservée)")
    print(f"\n  Les lignes auto-OCR sont surlignées en VERT dans le fichier.")
    print(f"  Le cabinet n'a plus qu'à valider visuellement avant de relancer Mizan.")


if __name__ == "__main__":
    main()
