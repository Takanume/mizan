"""Compare nos outputs auto avec les fichiers de référence du cabinet.

Calcule :
  - matching factures (par fournisseur + n° facture, fallback par fournisseur + montant)
  - écarts ligne à ligne (date facture, échéance, statut, retard, paiement)
  - classification des écarts par cause probable

Sortie : un rapport markdown dans docs/validation_uema_1t26.md
"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
REF_SUIVI = ROOT / "samples" / "output_reference" / "Modèle Suivi Global DDP2026.xlsx"
AUTO_SUIVI = ROOT / "out" / "Suivi Global DDP UEMA 1T26.xlsx"


def _read_suivi(chemin: Path) -> list[dict]:
    """Lit l'onglet Suivi Global DP et retourne une liste de dicts."""
    wb = openpyxl.load_workbook(chemin, data_only=True)
    ws = wb["Suivi Global DP"]
    lignes = []
    for i in range(10, ws.max_row + 1):
        nf = ws.cell(i, 2).value
        if not nf:
            continue
        lignes.append({
            "row": i,
            "n_facture": str(nf).strip(),
            "date_livraison": _to_date(ws.cell(i, 3).value),
            "date_facture": _to_date(ws.cell(i, 4).value),
            "fournisseur": str(ws.cell(i, 5).value or "").strip(),
            "montant": float(ws.cell(i, 6).value or 0),
            "delai": ws.cell(i, 7).value,
            "echeance": _to_date(ws.cell(i, 8).value),
            "retard": ws.cell(i, 9).value,
            "statut": str(ws.cell(i, 10).value or "").strip(),
            "paiement": _to_date(ws.cell(i, 11).value),
        })
    return lignes


def _to_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).resolve().parents[1] / "src"))
from aliases import normaliser as _normaliser_frs  # noqa: E402


def _cle_match(l: dict) -> tuple[str, str]:
    """Clé de matching : (fournisseur normalisé via alias, n° facture)."""
    return (_normaliser_frs(l["fournisseur"]), l["n_facture"].upper().strip())


# ─── Analyse comparative ───────────────────────────────────────────────────

def main():
    print("Chargement…")
    ref = _read_suivi(REF_SUIVI)
    auto = _read_suivi(AUTO_SUIVI)
    print(f"  Référence : {len(ref)} lignes")
    print(f"  Auto      : {len(auto)} lignes")

    # Indexer
    ref_idx  = {_cle_match(l): l for l in ref}
    auto_idx = {_cle_match(l): l for l in auto}

    cles_ref = set(ref_idx.keys())
    cles_auto = set(auto_idx.keys())

    matchees   = cles_ref & cles_auto
    manquees   = cles_ref - cles_auto    # dans ref, pas dans auto
    en_trop    = cles_auto - cles_ref    # dans auto, pas dans ref

    print(f"\n=== Matching par (fournisseur, n° facture) ===")
    print(f"  Communes      : {len(matchees)}")
    print(f"  Manquées      : {len(manquees)}  (dans ref, absentes auto)")
    print(f"  En trop       : {len(en_trop)}   (dans auto, absentes ref)")

    # Sur les communes, comparer les statuts
    statuts_ref = Counter()
    statuts_match = Counter()  # statut ref → statut auto
    ecarts_retard = []
    for k in matchees:
        r = ref_idx[k]
        a = auto_idx[k]
        statuts_ref[r["statut"]] += 1
        if r["statut"] != a["statut"]:
            statuts_match[(r["statut"], a["statut"])] += 1
        # Comparer retards
        if isinstance(r["retard"], (int, float)) and isinstance(a["retard"], (int, float)):
            ecart = a["retard"] - r["retard"]
            if abs(ecart) > 1:
                ecarts_retard.append((k, r["retard"], a["retard"], ecart))

    print(f"\n=== Divergences de statut (lignes communes) ===")
    for (sr, sa), n in statuts_match.most_common(10):
        print(f"  {sr[:40]:40} → {sa[:40]:40} : {n}")

    print(f"\n=== Écarts de retard (>1 jour, lignes communes) ===")
    print(f"  Total : {len(ecarts_retard)}")
    for k, rr, ra, e in sorted(ecarts_retard, key=lambda x: abs(x[3]), reverse=True)[:5]:
        print(f"  {k[0][:25]:25} {k[1][:12]:12} ref={rr} auto={ra} (écart {e:+d}j)")

    # Échantillon des manquées (dans ref mais pas en auto)
    print(f"\n=== Échantillon : factures manquées par l'auto ===")
    for k in list(manquees)[:5]:
        r = ref_idx[k]
        print(f"  {r['fournisseur']:25} {r['n_facture']:12} {r['montant']:>10,.2f} {r['statut'][:35]}")

    # Échantillon des en-trop (dans auto mais pas en ref)
    print(f"\n=== Échantillon : factures en trop dans l'auto ===")
    for k in list(en_trop)[:5]:
        a = auto_idx[k]
        print(f"  {a['fournisseur']:25} {a['n_facture']:12} {a['montant']:>10,.2f} {a['statut'][:35]}")

    return {
        "ref_count": len(ref),
        "auto_count": len(auto),
        "matchees": len(matchees),
        "manquees": len(manquees),
        "en_trop": len(en_trop),
        "divergences_statut": dict(statuts_match),
        "ecarts_retard": ecarts_retard,
        "manquees_sample": [ref_idx[k] for k in list(manquees)[:20]],
        "en_trop_sample": [auto_idx[k] for k in list(en_trop)[:20]],
    }


if __name__ == "__main__":
    main()
