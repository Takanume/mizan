"""Génère un Suivi Global côte à côte (cabinet vs auto).

Produit `out/Suivi Global - Comparaison Cabinet.xlsx` avec :
  • Bloc gauche  (10 colonnes)  : valeurs cabinet
  • Bloc droite  (10 colonnes)  : valeurs auto (Mizan)
  • Colonne finale              : observation (cause de l'écart si différent)
  • Cellules qui diffèrent      : fond jaune
  • Ligne entièrement cabinet   : fond bleu sur le bloc cabinet
  • Ligne entièrement auto      : fond rouge sur le bloc auto

Tri : dans l'ordre du fichier cabinet (les "en trop" auto à la fin).

Usage :
    python3 scripts/comparaison_visuelle.py
"""

from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
from aliases import normaliser  # noqa: E402

REF_SUIVI  = ROOT / "samples" / "output_reference" / "Modèle Suivi Global DDP2026.xlsx"
AUTO_SUIVI = ROOT / "out" / "Suivi Global DDP UEMA 1T26.xlsx"
SORTIE     = ROOT / "out" / "Suivi Global - Comparaison Cabinet.xlsx"

# ─── Couleurs ─────────────────────────────────────────────────────────────
NAVY      = "FF2B4C6F"
TEAL      = "FF4ECDC4"
PAPER     = "FFF8FAFC"
LINE_COL  = "FFD8E0EA"

FILL_HEADER_CAB = PatternFill("solid", fgColor="FF1E3A5F")   # bleu navy foncé
FILL_HEADER_AUTO = PatternFill("solid", fgColor="FF1E8C84")  # teal foncé
FILL_HEADER_OBS = PatternFill("solid", fgColor="FF142A44")
FILL_BAND_CAB   = PatternFill("solid", fgColor="FFEFF6FF")   # bleu très clair
FILL_BAND_AUTO  = PatternFill("solid", fgColor="FFECFDF5")   # vert/teal très clair
FILL_DIFF       = PatternFill("solid", fgColor="FFFEF3C7")   # jaune (cellule qui diffère)
FILL_MANQUEE    = PatternFill("solid", fgColor="FFDBEAFE")   # bleu (cabinet only)
FILL_EN_TROP    = PatternFill("solid", fgColor="FFFEE2E2")   # rouge (auto only)
FILL_OK         = PatternFill("solid", fgColor="FFD1FAE5")   # vert (ligne parfaite)

THIN = Side(border_style="thin", color=LINE_COL)
MEDIUM = Side(border_style="medium", color="FF1E3A5F")
BORDER_CELL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ONGLET = "Suivi Global DP"
LIGNE_DEBUT = 10

# Colonnes du Suivi Global (1 à 12 dans le fichier source)
LIBELLE_COLONNES = [
    "N° Facture",
    "Date Livraison",
    "Date Facture",
    "Fournisseur",
    "Montant TTC",
    "Délai",
    "Échéance",
    "Retard",
    "Statut",
    "Paiement",
]

# Clés correspondantes dans nos dicts
CLES = [
    "n_facture",
    "date_livraison",
    "date_facture",
    "fournisseur",
    "montant",
    "delai",
    "echeance",
    "retard",
    "statut",
    "paiement",
]


# ─── Lecture ──────────────────────────────────────────────────────────────

def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def lire_suivi(chemin: Path) -> list[dict]:
    wb = openpyxl.load_workbook(chemin, data_only=True)
    ws = wb[ONGLET]
    lignes = []
    for r in range(LIGNE_DEBUT, ws.max_row + 1):
        nf = ws.cell(r, 2).value
        if not nf:
            continue
        lignes.append({
            "n_facture":     str(nf).strip(),
            "date_livraison":_to_date(ws.cell(r, 3).value),
            "date_facture":  _to_date(ws.cell(r, 4).value),
            "fournisseur":   str(ws.cell(r, 5).value or "").strip(),
            "montant":       float(ws.cell(r, 6).value or 0),
            "delai":         ws.cell(r, 7).value,
            "echeance":      _to_date(ws.cell(r, 8).value),
            "retard":        ws.cell(r, 9).value,
            "statut":        str(ws.cell(r, 10).value or "").strip(),
            "paiement":      _to_date(ws.cell(r, 11).value),
        })
    wb.close()
    return lignes


def cle(li: dict) -> tuple[str, str]:
    return (normaliser(li["fournisseur"]), str(li["n_facture"]).upper().strip())


# ─── Calcul des écarts ────────────────────────────────────────────────────

def diffs(cab: dict, auto: dict) -> tuple[set[str], list[str]]:
    """Retourne (clés qui diffèrent, liste lisible des différences)."""
    differentes: set[str] = set()
    raisons: list[str] = []
    for k in CLES:
        v_cab, v_auto = cab.get(k), auto.get(k)
        if k == "montant":
            if abs(float(v_cab or 0) - float(v_auto or 0)) >= 0.01:
                differentes.add(k)
        elif v_cab != v_auto:
            differentes.add(k)
    if "statut" in differentes:
        raisons.append(f"statut : {cab['statut']} → {auto['statut']}")
    if "echeance" in differentes:
        if cab["echeance"] and auto["echeance"]:
            raisons.append(f"échéance Δ={(auto['echeance']-cab['echeance']).days:+d}j")
        else:
            raisons.append("échéance absente d'un côté")
    if "paiement" in differentes:
        if cab["paiement"] and auto["paiement"]:
            raisons.append(f"paiement Δ={(auto['paiement']-cab['paiement']).days:+d}j")
        elif cab["paiement"]:
            raisons.append("paiement manquant auto")
        else:
            raisons.append("paiement ajouté auto")
    if "delai" in differentes:
        raisons.append(f"délai {cab['delai']}→{auto['delai']}")
    if "date_facture" in differentes:
        raisons.append("date facture diffère")
    if "fournisseur" in differentes:
        raisons.append(f"nom : « {cab['fournisseur']} » vs « {auto['fournisseur']} »")
    return differentes, raisons


# ─── Génération ───────────────────────────────────────────────────────────

def generer():
    if not REF_SUIVI.exists():
        sys.exit(f"❌ Référence introuvable : {REF_SUIVI}")
    if not AUTO_SUIVI.exists():
        sys.exit(f"❌ Auto introuvable : {AUTO_SUIVI}")

    print("→ Lecture des deux fichiers…")
    ref  = lire_suivi(REF_SUIVI)
    auto = lire_suivi(AUTO_SUIVI)
    auto_idx = {cle(l): l for l in auto}
    ref_keys = {cle(l) for l in ref}
    print(f"  cabinet={len(ref)} | auto={len(auto)}")

    # ─── Création du classeur ────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparaison côte à côte"

    n_col_cab = len(LIBELLE_COLONNES)        # 10
    n_col_auto = len(LIBELLE_COLONNES)       # 10
    # Layout : col 1 = #, cols 2-11 = cabinet, cols 12-21 = auto, col 22 = observation
    col_cab_start  = 2
    col_auto_start = col_cab_start + n_col_cab     # 12
    col_obs        = col_auto_start + n_col_auto   # 22

    # Titre
    ws.cell(1, 1).value = "MIZAN — COMPARAISON SUIVI GLOBAL  ·  CABINET (gauche) vs AUTO (droite)"
    ws.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_obs)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # En-tête de section (ligne 3)
    ws.cell(3, col_cab_start).value = "VALEURS CABINET (référence)"
    ws.cell(3, col_cab_start).font = Font(bold=True, color="FFFFFFFF", size=11)
    ws.cell(3, col_cab_start).fill = FILL_HEADER_CAB
    ws.cell(3, col_cab_start).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=3, start_column=col_cab_start,
                   end_row=3, end_column=col_cab_start + n_col_cab - 1)

    ws.cell(3, col_auto_start).value = "VALEURS AUTO (Mizan)"
    ws.cell(3, col_auto_start).font = Font(bold=True, color="FFFFFFFF", size=11)
    ws.cell(3, col_auto_start).fill = FILL_HEADER_AUTO
    ws.cell(3, col_auto_start).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=3, start_column=col_auto_start,
                   end_row=3, end_column=col_auto_start + n_col_auto - 1)

    ws.cell(3, col_obs).value = "OBSERVATION"
    ws.cell(3, col_obs).font = Font(bold=True, color="FFFFFFFF", size=11)
    ws.cell(3, col_obs).fill = FILL_HEADER_OBS
    ws.cell(3, col_obs).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24

    # Ligne d'en-tête de colonnes (ligne 4)
    ws.cell(4, 1).value = "#"
    ws.cell(4, 1).font = Font(bold=True, color="FFFFFFFF")
    ws.cell(4, 1).fill = FILL_HEADER_OBS
    ws.cell(4, 1).alignment = Alignment(horizontal="center", vertical="center")
    for i, lib in enumerate(LIBELLE_COLONNES):
        c1 = ws.cell(4, col_cab_start + i, lib)
        c1.font = Font(bold=True, color="FFFFFFFF", size=10)
        c1.fill = FILL_HEADER_CAB
        c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c2 = ws.cell(4, col_auto_start + i, lib)
        c2.font = Font(bold=True, color="FFFFFFFF", size=10)
        c2.fill = FILL_HEADER_AUTO
        c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    # Figer en-têtes
    ws.freeze_panes = ws.cell(5, col_cab_start)

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 5
    largeurs = [13, 12, 12, 28, 12, 7, 12, 8, 28, 12]
    for i, w in enumerate(largeurs):
        ws.column_dimensions[get_column_letter(col_cab_start + i)].width = w
        ws.column_dimensions[get_column_letter(col_auto_start + i)].width = w
    ws.column_dimensions[get_column_letter(col_obs)].width = 60

    # ─── Écriture des lignes ─────────────────────────────────────────
    def ecrire_bloc(row, col_start, ligne, fill_band, fill_diff_keys=None, vide=False):
        """Écrit 10 colonnes pour une ligne (cabinet ou auto)."""
        for i, k in enumerate(CLES):
            col = col_start + i
            cell = ws.cell(row, col)
            if not vide and ligne is not None:
                cell.value = ligne.get(k)
            cell.fill = fill_diff if (fill_diff_keys and k in fill_diff_keys) else fill_band
            cell.border = BORDER_CELL
            cell.font = Font(size=10)
            # Format
            if k in ("date_livraison", "date_facture", "echeance", "paiement"):
                cell.number_format = "DD/MM/YYYY"
            elif k == "montant":
                cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="left" if k in ("fournisseur", "statut") else "center",
                                       vertical="center")

    fill_diff = FILL_DIFF

    row = 5
    n_ok = n_diff = n_manq = n_trop = 0

    # 1. Lignes communes et manquées — dans l'ordre du cabinet
    for cab_ligne in ref:
        k = cle(cab_ligne)
        auto_ligne = auto_idx.get(k)
        ws.cell(row, 1).value = row - 4
        ws.cell(row, 1).font = Font(bold=True, size=9, color="FF64748B")
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 1).border = BORDER_CELL

        if auto_ligne is None:
            # Manquée — cabinet seul
            ecrire_bloc(row, col_cab_start, cab_ligne, FILL_MANQUEE)
            ecrire_bloc(row, col_auto_start, None, FILL_BAND_AUTO, vide=True)
            ws.cell(row, col_obs).value = "⚠ Ligne du cabinet absente de notre output"
            ws.cell(row, col_obs).fill = FILL_MANQUEE
            ws.cell(row, col_obs).font = Font(size=10, bold=True, color="FF1E3A5F")
            n_manq += 1
        else:
            differs, raisons = diffs(cab_ligne, auto_ligne)
            if not differs:
                ecrire_bloc(row, col_cab_start, cab_ligne, FILL_OK)
                ecrire_bloc(row, col_auto_start, auto_ligne, FILL_OK)
                ws.cell(row, col_obs).value = "✓ identique"
                ws.cell(row, col_obs).fill = FILL_OK
                ws.cell(row, col_obs).font = Font(size=10, color="FF2E864E")
                n_ok += 1
            else:
                ecrire_bloc(row, col_cab_start, cab_ligne, FILL_BAND_CAB, differs)
                ecrire_bloc(row, col_auto_start, auto_ligne, FILL_BAND_AUTO, differs)
                ws.cell(row, col_obs).value = " · ".join(raisons)
                ws.cell(row, col_obs).fill = FILL_DIFF
                ws.cell(row, col_obs).font = Font(size=10)
                n_diff += 1
        ws.cell(row, col_obs).border = BORDER_CELL
        ws.cell(row, col_obs).alignment = Alignment(horizontal="left", vertical="center",
                                                     wrap_text=True)
        row += 1

    # 2. En trop — auto seul
    for auto_ligne in auto:
        if cle(auto_ligne) in ref_keys:
            continue
        ws.cell(row, 1).value = row - 4
        ws.cell(row, 1).font = Font(bold=True, size=9, color="FF64748B")
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 1).border = BORDER_CELL
        ecrire_bloc(row, col_cab_start, None, FILL_BAND_CAB, vide=True)
        ecrire_bloc(row, col_auto_start, auto_ligne, FILL_EN_TROP)
        ws.cell(row, col_obs).value = "⚠ Ligne en trop — absente de la déclaration cabinet"
        ws.cell(row, col_obs).fill = FILL_EN_TROP
        ws.cell(row, col_obs).font = Font(size=10, bold=True, color="FFB91C1C")
        ws.cell(row, col_obs).border = BORDER_CELL
        ws.cell(row, col_obs).alignment = Alignment(horizontal="left", vertical="center",
                                                     wrap_text=True)
        n_trop += 1
        row += 1

    # ─── Onglet Légende ──────────────────────────────────────────────
    leg = wb.create_sheet("Légende", 1)
    leg["A1"] = "Légende — Comparaison côte à côte"
    leg["A1"].font = Font(bold=True, size=14, color="FFFFFFFF")
    leg["A1"].fill = PatternFill("solid", fgColor=NAVY)
    leg.merge_cells("A1:D1")
    leg["A1"].alignment = Alignment(horizontal="center")

    leg["A3"], leg["B3"], leg["C3"] = "Couleur", "Nb lignes", "Description"
    for c in ("A3", "B3", "C3"):
        leg[c].font = Font(bold=True)
        leg[c].fill = PatternFill("solid", fgColor=PAPER)
    rows_leg = [
        ("Vert",  FILL_OK,       n_ok,   "Ligne identique au cabinet (toutes colonnes alignées)"),
        ("Jaune", FILL_DIFF,     n_diff, "Ligne matchée mais au moins une cellule diffère (cellule jaune)"),
        ("Bleu",  FILL_MANQUEE,  n_manq, "Présente cabinet, absente auto (bloc auto vide)"),
        ("Rouge", FILL_EN_TROP,  n_trop, "Présente auto, absente cabinet (bloc cabinet vide)"),
    ]
    for i, (label, fill, n, desc) in enumerate(rows_leg, start=4):
        leg.cell(i, 1, label).fill = fill
        leg.cell(i, 2, n).font = Font(bold=True)
        leg.cell(i, 3, desc)
    leg.column_dimensions["A"].width = 10
    leg.column_dimensions["B"].width = 12
    leg.column_dimensions["C"].width = 80

    leg["A10"] = "Lecture des cellules"
    leg["A10"].font = Font(bold=True, size=12)
    leg["A12"] = "Jaune sur une cellule = la valeur cabinet et auto diffèrent sur cette colonne précisément."
    leg["A13"] = "Bloc cabinet en bleu  = ligne uniquement présente côté cabinet."
    leg["A14"] = "Bloc auto en rouge    = ligne uniquement présente côté auto."

    wb.save(SORTIE)
    print(f"\n✓ Fichier généré : {SORTIE}")
    print(f"  🟢 Vert  (identiques) : {n_ok:>4}")
    print(f"  🟡 Jaune (écarts)     : {n_diff:>4}")
    print(f"  🔵 Bleu  (manquées)   : {n_manq:>4}")
    print(f"  🔴 Rouge (en trop)    : {n_trop:>4}")
    print(f"  Total                 : {n_ok + n_diff + n_manq + n_trop}")


if __name__ == "__main__":
    generer()
