"""Rend un fichier Suivi Global cabinet compatible Google Sheets.

Le template du cabinet utilise des fonctionnalités Excel non portables vers Sheets :
  - Tableau1[[#This Row],[Fournisseur]]  → références de tableau structuré
  - SuiviFrs                              → plage nommée

Ce script réécrit les formules des colonnes G/H/I/J en syntaxe portable :
  G (Délai)     : =IF(E10="","",IFERROR(VLOOKUP(E10,'Base Frs Permanente'!$B:$C,2,FALSE),"Vérifier..."))
  H (Échéance)  : =IF(E10="","",C10+G10-1)
  I (Retard)    : =IF(E10="","",IF(H10-K10>0," ",H10-K10))
  J (Statut)    : =IF(E10="","",IF(K10="","Non encore payée",IF(I10<0,"Attention, paiement hors délais","OK RAS")))

Le fichier source est conservé intact ; le résultat est sauvé sous un nouveau nom.

Usage :
    python3 scripts/patch_template_sheets.py [chemin_source] [chemin_destination]

Par défaut, lit `samples/output_reference/Modèle Suivi Global DDP2026.xlsx`
et écrit `out/Modèle Suivi Global DDP2026 - Sheets.xlsx`.
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl


ONGLET = "Suivi Global DP"
LIGNE_DEBUT = 10
LIGNE_FIN_MAX = 5000  # marge confortable

COL_FOURNISSEUR = "E"   # référence métier
COL_DATE_LIV    = "C"
COL_DELAI       = "G"
COL_ECHEANCE    = "H"
COL_RETARD      = "I"
COL_STATUT      = "J"
COL_PAIEMENT    = "K"

# Plage de la base fournisseurs (col B = nom, col C = délai)
PLAGE_BASE = "'Base Frs Permanente'!$B$4:$C$2000"


def formule_delai(row: int) -> str:
    return (
        f'=IF({COL_FOURNISSEUR}{row}="","",IFERROR(VLOOKUP({COL_FOURNISSEUR}{row},'
        f'{PLAGE_BASE},2,FALSE),"Vérifier la syntaxe : Fournisseur introuvable au niveau du suivi des conventions"))'
    )


def formule_echeance(row: int) -> str:
    return f'=IF({COL_FOURNISSEUR}{row}="","",{COL_DATE_LIV}{row}+{COL_DELAI}{row}-1)'


def formule_retard(row: int) -> str:
    return (
        f'=IF({COL_FOURNISSEUR}{row}="","",'
        f'IF({COL_ECHEANCE}{row}-{COL_PAIEMENT}{row}>0," ",'
        f'{COL_ECHEANCE}{row}-{COL_PAIEMENT}{row}))'
    )


def formule_statut(row: int) -> str:
    return (
        f'=IF({COL_FOURNISSEUR}{row}="","",'
        f'IF({COL_PAIEMENT}{row}="","Non encore payée",'
        f'IF({COL_RETARD}{row}<0,"Attention, paiement hors délais","OK RAS")))'
    )


def cellule_contient_formule_tableau(value) -> bool:
    """True si la cellule contient une formule avec syntaxe non-portable."""
    if not isinstance(value, str):
        return False
    if not value.startswith("="):
        return False
    return ("Tableau1[" in value) or ("SuiviFrs" in value)


def patcher_fichier(chemin_src: Path, chemin_dst: Path) -> dict[str, int]:
    """Réécrit les formules non portables dans un fichier Suivi Global.

    Retourne un compteur des cellules patchées par colonne.
    """
    chemin_dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(chemin_src, chemin_dst)
    wb = openpyxl.load_workbook(chemin_dst)
    if ONGLET not in wb.sheetnames:
        raise ValueError(f"Onglet '{ONGLET}' introuvable")
    ws = wb[ONGLET]

    compteur = {"G": 0, "H": 0, "I": 0, "J": 0}
    lignes_patchees = 0

    last_row = min(ws.max_row, LIGNE_FIN_MAX)
    for row in range(LIGNE_DEBUT, last_row + 1):
        ligne_modifiee = False
        # G — Délai
        if cellule_contient_formule_tableau(ws[f"G{row}"].value):
            ws[f"G{row}"] = formule_delai(row)
            compteur["G"] += 1
            ligne_modifiee = True
        # H — Échéance
        if cellule_contient_formule_tableau(ws[f"H{row}"].value):
            ws[f"H{row}"] = formule_echeance(row)
            compteur["H"] += 1
            ligne_modifiee = True
        # I — Retard
        if cellule_contient_formule_tableau(ws[f"I{row}"].value):
            ws[f"I{row}"] = formule_retard(row)
            compteur["I"] += 1
            ligne_modifiee = True
        # J — Statut
        if cellule_contient_formule_tableau(ws[f"J{row}"].value):
            ws[f"J{row}"] = formule_statut(row)
            compteur["J"] += 1
            ligne_modifiee = True
        if ligne_modifiee:
            lignes_patchees += 1

    # Supprimer le tableau structuré Tableau1 (Sheets ne le comprend pas non plus)
    if "Tableau1" in ws.tables:
        del ws.tables["Tableau1"]

    # Supprimer la plage nommée SuiviFrs (remplacée par référence directe)
    noms_a_supprimer = ["SuiviFrs", "BaseFournisseurs", "base", "basefrs", "pc", "xxx"]
    for n in noms_a_supprimer:
        if n in wb.defined_names:
            del wb.defined_names[n]

    wb.save(chemin_dst)
    print(f"✓ {lignes_patchees} lignes patchées")
    return compteur


def main():
    racine = Path(__file__).resolve().parents[1]
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else (
        racine / "samples" / "output_reference" / "Modèle Suivi Global DDP2026.xlsx"
    )
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else (
        racine / "out" / "Modèle Suivi Global DDP2026 - Sheets.xlsx"
    )
    print(f"Source : {src}")
    print(f"Destination : {dst}")
    compteur = patcher_fichier(src, dst)
    print(f"\nFormules réécrites par colonne :")
    for col, n in compteur.items():
        print(f"  {col} : {n}")


if __name__ == "__main__":
    main()
