"""Génère un référentiel DGI fournisseurs pré-rempli (seed initial).

Croise :
  - La Base Frs Permanente existante (code + nom + délai)
  - Les données DGI extraites du Simpl de référence (IF, ICE, RC, adresse, ville)

Produit un fichier Excel structuré :
  samples/input/Référentiel DGI - Cabinet.xlsx

Ce fichier devient la source unique du cabinet — un fournisseur, une ligne.
Le cabinet pourra ensuite l'enrichir incrémentalement (ajouter RC, adresses…).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
REF_SUIVI = ROOT / "samples" / "output_reference" / "Modèle Suivi Global DDP2026.xlsx"
REF_SIMPL = ROOT / "samples" / "output_reference" / "Simpl délais de paiements UEMA -  TR 01-2026.xlsx"
SORTIE    = ROOT / "samples" / "input" / "Référentiel DGI - Cabinet.xlsx"


# ─── Étape 1 — Charger la base actuelle (code → nom, délai) ───────────────

def charger_base_actuelle() -> dict[str, dict]:
    """Charge la base Frs Permanente avec code/nom/délai/observations.

    Gère les doublons de codes (le fichier cabinet en contient plusieurs)
    en gardant la PREMIÈRE occurrence — les suivantes sont supposées être
    des variantes / typos ajoutés ensuite. Les doublons avec délais
    différents sont signalés.
    """
    wb = openpyxl.load_workbook(REF_SUIVI, data_only=True)
    ws = wb["Base Frs Permanente"]
    base: dict[str, dict] = {}
    doublons: dict[str, list[dict]] = {}

    for i in range(4, ws.max_row + 1):
        code = ws.cell(i, 1).value
        nom  = ws.cell(i, 2).value
        delai = ws.cell(i, 3).value
        obs  = ws.cell(i, 4).value
        if not code or not isinstance(code, str) or not code.startswith("FRS"):
            continue
        code = code.strip()
        entry = {
            "row": i,
            "nom": str(nom).strip() if nom else "",
            "delai": int(delai) if delai is not None else 60,
            "observations": str(obs).strip() if obs else "",
        }
        if code in base:
            # Doublon : on garde la première, on collectionne pour alerter
            doublons.setdefault(code, [base[code].copy() | {"row": "?"}]).append(entry)
        else:
            base[code] = entry

    # Trace les doublons (utile pour le rapport)
    if doublons:
        print(f"\n⚠ {len(doublons)} code(s) dupliqué(s) dans le fichier cabinet :")
        for code, lignes in doublons.items():
            delais = {l["delai"] for l in lignes}
            flag = " 🚨 DÉLAIS DIFFÉRENTS" if len(delais) > 1 else ""
            print(f"  {code}{flag}")
            for l in lignes:
                marqueur = "  ← retenu" if l is lignes[0] else "  ← ignoré"
                print(f"    R{l['row']!s:>3} | {l['nom']:40} | délai={l['delai']:>4}{marqueur}")
        print()

    wb.close()
    return base


# ─── Étape 2 — Charger les données DGI du Simpl de référence ───────────────

def charger_dgi_simpl() -> dict[str, dict]:
    """Construit un dict {nom_normalisé → {if, ice, rc, adresse, ville}}."""
    wb = openpyxl.load_workbook(REF_SIMPL, data_only=True)
    ws = wb.active
    dgi = {}
    for i in range(20, ws.max_row + 1):
        nom = ws.cell(i, 3).value
        if not nom:
            continue
        key = str(nom).strip().upper()
        if key in dgi:
            continue
        dgi[key] = {
            "if":      str(ws.cell(i, 1).value).strip() if ws.cell(i, 1).value else "",
            "ice":     str(ws.cell(i, 2).value).strip() if ws.cell(i, 2).value else "",
            "rc":      str(ws.cell(i, 4).value).strip() if ws.cell(i, 4).value else "",
            "adresse": str(ws.cell(i, 5).value).strip() if ws.cell(i, 5).value else "",
            "ville":   str(ws.cell(i, 6).value).strip() if ws.cell(i, 6).value else "",
        }
    wb.close()
    return dgi


# ─── Étape 3 — Créer le fichier référentiel ────────────────────────────────

NAVY = "FF2B4C6F"
TEAL = "FF4ECDC4"
PAPER = "FFF8FAFC"
LINE_COLOR = "FFD8E0EA"

HEADERS = [
    ("N° Fournisseur",      14),  # A
    ("Nom Fournisseur",     38),  # B
    ("Délai (jours)",       14),  # C
    ("N° IF",               14),  # D
    ("N° ICE",              22),  # E
    ("N° RC",               12),  # F
    ("Adresse siège social",40),  # G
    ("Ville",               18),  # H
    ("Secteur d'activité",  24),  # I
    ("Nature marchandises", 30),  # J
    ("Observations",        24),  # K
]


def creer_referentiel():
    base = charger_base_actuelle()
    dgi  = charger_dgi_simpl()

    print(f"Base actuelle : {len(base)} fournisseurs")
    print(f"DGI Simpl     : {len(dgi)} fournisseurs")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Référentiel DGI"

    # En-tête : titre
    ws["A1"] = "RÉFÉRENTIEL DGI FOURNISSEURS — NEXTOR / MIZAN"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Note d'aide
    ws["A2"] = ("Le cabinet enrichit cette base au fil des trimestres. "
                "Les colonnes DGI (IF, ICE, RC, adresse, ville) servent à pré-remplir le formulaire Simpl.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    ws["A2"].font = Font(italic=True, color="FF64748B")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # Ligne d'en-tête de colonnes
    header_row = 3
    for col_idx, (label, width) in enumerate(HEADERS, start=1):
        c = ws.cell(header_row, col_idx, label)
        c.font = Font(bold=True, color="FFFFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = width
    ws.row_dimensions[header_row].height = 30

    # Borders pour les data
    thin = Side(border_style="thin", color=LINE_COLOR)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Tri : d'abord ceux qui ont des données DGI, puis le reste
    enrichis = []
    autres = []
    for code, info in base.items():
        nom_key = info["nom"].upper()
        if nom_key in dgi:
            enrichis.append((code, info, dgi[nom_key]))
        else:
            autres.append((code, info, {}))

    # Tri alphabétique par nom dans chaque groupe
    enrichis.sort(key=lambda x: x[1]["nom"])
    autres.sort(key=lambda x: x[1]["nom"])

    nb_enrichis = 0
    row = header_row + 1
    for code, info, dgi_info in enrichis + autres:
        ws.cell(row, 1, code)
        ws.cell(row, 2, info["nom"])
        ws.cell(row, 3, info["delai"])
        ws.cell(row, 4, dgi_info.get("if", ""))
        ws.cell(row, 5, dgi_info.get("ice", ""))
        ws.cell(row, 6, dgi_info.get("rc", ""))
        ws.cell(row, 7, dgi_info.get("adresse", ""))
        ws.cell(row, 8, dgi_info.get("ville", ""))
        ws.cell(row, 9, "")   # secteur — à compléter par le cabinet
        ws.cell(row, 10, "")  # nature marchandises
        ws.cell(row, 11, info.get("observations", ""))

        # Style ligne
        bg = "FFE9F5F3" if dgi_info.get("if") else (PAPER if (row - header_row) % 2 == 0 else "FFFFFFFF")
        for col_idx in range(1, len(HEADERS) + 1):
            c = ws.cell(row, col_idx)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = cell_border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx == 3:  # délai
                c.alignment = Alignment(horizontal="center", vertical="center")

        if dgi_info.get("if"):
            nb_enrichis += 1
        row += 1

    # Figer les volets pour navigation aisée
    ws.freeze_panes = ws.cell(header_row + 1, 3)

    # Ajouter une légende en haut à droite (zone libre)
    legende_row = 2
    ws.cell(legende_row, len(HEADERS) + 2, "Légende :")
    # (legende compacte : pas faisable proprement sans complexifier — on saute)

    SORTIE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SORTIE)
    print(f"\n✓ Référentiel généré : {SORTIE}")
    print(f"  Total fournisseurs : {row - header_row - 1}")
    print(f"  Enrichis (avec IF/ICE) : {nb_enrichis}  ({nb_enrichis/(row-header_row-1)*100:.0f} %)")
    print(f"  À compléter par le cabinet : {(row-header_row-1) - nb_enrichis}")


if __name__ == "__main__":
    creer_referentiel()
