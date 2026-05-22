"""Surcharge manuelle pour une facture identifiée (résout D-006 et D-003).

Permet au cabinet de fournir :
  - la **vraie date** d'une facture (D-006) quand elle ne figure pas dans Sage
    (cas des factures reportées AN où Sage force la date au 01/01) ;
  - un **délai spécifique** (D-003) qui surcharge celui du fournisseur
    (cas où la même facture est régie par une convention ponctuelle —
    ex : PPRIME 60 j sur produits standards mais 90 j sur prestations).

Format du fichier de corrections (Excel) :

  Onglet `Corrections Dates` :
    A : N° Fournisseur  (FRSxxxxxxx, optionnel — pour désambiguïsation)
    B : Nom Fournisseur (clé secondaire de matching)
    C : N° Facture
    D : Date Facture réelle      (vraie date, depuis le PDF)
    E : Date Livraison           (si différente, sinon = Date Facture)
    F : Observations             (libre)
    G : Délai spécifique (j)     (optionnel — surcharge le délai fournisseur)

Le matching est fait sur `(N° Fournisseur, N° Facture)` en priorité, puis
sur `(Nom Fournisseur, N° Facture)` en fallback.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from models import Facture  # noqa: E402


ONGLET_CORRECTIONS = "Corrections Dates"


@dataclass(frozen=True)
class CorrectionDate:
    """Une surcharge pour une facture identifiée (date et/ou délai).

    Note : malgré le nom historique « CorrectionDate », l'objet porte aussi
    une éventuelle surcharge de délai (`delai_jours`) — c'est une extension
    Phase 2 Sprint 3 pour D-003.
    """
    code_fournisseur: Optional[str]            # ex. "FRS0000235" (peut être None si non précisé)
    nom_fournisseur: str                       # ex. "PPRIME"
    n_facture: str                             # ex. "25/023"
    date_facture: date                         # vraie date d'émission
    date_livraison: Optional[date] = None      # None → = date_facture
    observations: Optional[str] = None
    delai_jours: Optional[int] = None          # surcharge du délai pour CETTE facture


# ─── Loader ────────────────────────────────────────────────────────────────

def _str_or_none(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def charger_corrections(chemin: Path | str) -> list[CorrectionDate]:
    """Charge les corrections de date depuis un fichier Excel.

    Retourne une liste vide si le fichier ou l'onglet n'existe pas.
    """
    chemin = Path(chemin)
    if not chemin.exists():
        return []
    wb = openpyxl.load_workbook(chemin, data_only=True, read_only=True)
    try:
        if ONGLET_CORRECTIONS not in wb.sheetnames:
            return []
        ws = wb[ONGLET_CORRECTIONS]
        corrections: list[CorrectionDate] = []
        # Trouver la première ligne de données
        debut = 2
        for i in range(1, min(ws.max_row, 10) + 1):
            v = ws.cell(i, 3).value  # col C = N° Facture
            if isinstance(v, (str, int)) and v and not str(v).strip().lower().startswith("n°"):
                debut = i
                break

        for row in ws.iter_rows(min_row=debut, values_only=True):
            if not row or len(row) < 4:
                continue
            row = list(row) + [None] * (7 - len(row))
            code = _str_or_none(row[0])
            nom  = _str_or_none(row[1])
            n_fact = _str_or_none(row[2])
            d_fact = _to_date(row[3])
            d_livr = _to_date(row[4])
            obs    = _str_or_none(row[5])
            # Col G : délai spécifique en jours (Sprint 3 — D-003)
            delai = None
            if row[6] is not None:
                try:
                    delai = int(row[6])
                    if delai <= 0 or delai > 365:
                        delai = None
                except (ValueError, TypeError):
                    pass

            if not n_fact or d_fact is None or not nom:
                continue

            corrections.append(CorrectionDate(
                code_fournisseur=code,
                nom_fournisseur=nom,
                n_facture=n_fact,
                date_facture=d_fact,
                date_livraison=d_livr,
                observations=obs,
                delai_jours=delai,
            ))
    finally:
        wb.close()
    return corrections


# ─── Index pour matching efficace ──────────────────────────────────────────

def construire_index(corrections: list[CorrectionDate]) -> dict[tuple[str, str], CorrectionDate]:
    """Indexe par (code_fournisseur OU nom_normalisé, n° facture normalisé).

    Chaque correction génère 1 ou 2 entrées :
      - (code_fournisseur, n_fact)  si code fourni
      - (nom_normalisé,     n_fact)  toujours (fallback)
    """
    idx: dict[tuple[str, str], CorrectionDate] = {}
    for c in corrections:
        n_norm = c.n_facture.strip().upper()
        nom_norm = c.nom_fournisseur.strip().upper()
        if c.code_fournisseur:
            idx[(c.code_fournisseur.strip(), n_norm)] = c
        idx[(nom_norm, n_norm)] = c
    return idx


def trouver_correction(
    facture: Facture,
    index: dict[tuple[str, str], CorrectionDate],
) -> Optional[CorrectionDate]:
    """Trouve la correction applicable à une facture, ou None."""
    if not facture.n_facture:
        return None
    n_norm = facture.n_facture.strip().upper()
    # Match par code fournisseur en priorité
    c = index.get((facture.code_fournisseur, n_norm))
    if c:
        return c
    # Fallback : par nom
    return index.get((facture.nom_fournisseur.strip().upper(), n_norm))


# ─── Application des corrections ───────────────────────────────────────────

def appliquer_correction(
    facture: Facture,
    correction: CorrectionDate,
) -> Facture:
    """Retourne une nouvelle facture avec date et délai éventuellement surchargés."""
    from dataclasses import replace
    return replace(
        facture,
        date_facture=correction.date_facture,
        date_livraison=correction.date_livraison or correction.date_facture,
        delai_surcharge=correction.delai_jours,
    )
