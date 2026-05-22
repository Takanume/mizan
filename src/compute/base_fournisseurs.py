"""Chargement du référentiel fournisseurs.

Deux formats supportés :

1. **Base Frs Permanente** (legacy — 4 colonnes) :
     A : N° Fournisseur (FRSxxxxxxx)
     B : Nom Fournisseur
     C : Délai (jours)
     D : Observations

2. **Référentiel DGI Cabinet** (Phase 2 — 11 colonnes enrichies) :
     A : N° Fournisseur
     B : Nom Fournisseur
     C : Délai (jours)
     D : N° IF
     E : N° ICE
     F : N° RC
     G : Adresse siège social
     H : Ville
     I : Secteur d'activité
     J : Nature marchandises
     K : Observations

Le loader auto-détecte le format selon l'onglet présent.
Les champs DGI absents restent None.

Si un fournisseur du GL est absent de la base, on lui applique le
délai par défaut (60 jours).

Les délais non-standards (hors {30, 45, 60, 75, 90, 120}) sont conservés tels
quels et signalés via `stats_delais_anormaux()`. Le cabinet doit corriger lui-même
les vraies erreurs de saisie — Mizan n'invente pas de correction.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import openpyxl

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from models import Fournisseur  # noqa: E402


DELAI_PAR_DEFAUT_JOURS = 60

# Noms d'onglets reconnus, dans l'ordre de préférence
ONGLETS_CONNUS = (
    "Référentiel DGI",         # Phase 2 — enrichi
    "Base Frs Permanente",     # Legacy
)


def _str_or_none(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _detecter_onglet(wb) -> str:
    """Retourne le nom du premier onglet connu trouvé dans le classeur."""
    for nom in ONGLETS_CONNUS:
        if nom in wb.sheetnames:
            return nom
    raise ValueError(
        f"Aucun onglet reconnu dans le classeur. Attendu : {ONGLETS_CONNUS}, "
        f"présent : {wb.sheetnames}"
    )


def _trouver_ligne_donnees(ws) -> int:
    """Trouve la première ligne contenant un code fournisseur FRSxxx.

    Les fichiers ont parfois un titre + ligne d'aide + en-tête sur 1 à 3 lignes,
    donc on scanne jusqu'à trouver la première vraie donnée.
    """
    for i in range(1, min(ws.max_row, 10) + 1):
        v = ws.cell(i, 1).value
        if isinstance(v, str) and v.strip().startswith("FRS"):
            return i
    # Fallback : ligne 2 (ancienne convention)
    return 2


DELAIS_STANDARDS = (0, 30, 45, 60, 75, 90, 120)

# Délais non-standards relevés à la lecture du référentiel — à valider par le cabinet.
# Format : { code_fournisseur: (nom, délai, standard_proche) }
_delais_atypiques: dict[str, tuple[str, int, int]] = {}


def _signaler_delai(code: str, nom: str, delai: int) -> int:
    """Garde le délai tel quel, mais le note si non-standard."""
    if delai not in DELAIS_STANDARDS:
        proche = min(DELAIS_STANDARDS, key=lambda s: abs(s - delai))
        _delais_atypiques[code] = (nom, delai, proche)
    return delai


def stats_delais_anormaux() -> dict:
    """Retourne les délais non-standards relevés au chargement (sans correction)."""
    return {"atypiques": dict(_delais_atypiques)}


def charger_base_fournisseurs(
    chemin: Path | str,
    onglet: Optional[str] = None,
) -> dict[str, Fournisseur]:
    """Charge un référentiel fournisseurs (legacy 4 colonnes ou DGI 11 colonnes).

    Le format est auto-détecté. Si `onglet` est passé, on l'utilise directement
    sinon on cherche dans `ONGLETS_CONNUS`.

    Retourne : dict {code_fournisseur → Fournisseur}.
    Les lignes sans code valide sont ignorées.
    """
    chemin = Path(chemin)
    wb = openpyxl.load_workbook(chemin, data_only=True, read_only=True)
    try:
        if onglet is None:
            onglet = _detecter_onglet(wb)
        elif onglet not in wb.sheetnames:
            raise ValueError(f"Onglet '{onglet}' introuvable dans {chemin.name}")
        ws = wb[onglet]

        debut = _trouver_ligne_donnees(ws)
        base: dict[str, Fournisseur] = {}
        _delais_atypiques.clear()

        for row in ws.iter_rows(min_row=debut, values_only=True):
            if not row:
                continue
            row = list(row) + [None] * (11 - len(row))  # padding pour accès safe

            code = _str_or_none(row[0])
            if not code or not code.startswith("FRS"):
                continue
            if row[2] is None:
                continue

            nom = _str_or_none(row[1]) or ""
            delai = int(row[2])
            _signaler_delai(code, nom, delai)
            base[code] = Fournisseur(
                code=code,
                nom=nom,
                delai_convenu_jours=delai,
                # Champs DGI (uniquement présents en format enrichi)
                n_if=_str_or_none(row[3]),
                n_ice=_str_or_none(row[4]),
                n_rc=_str_or_none(row[5]),
                adresse=_str_or_none(row[6]),
                ville_rc=_str_or_none(row[7]),
                secteur_activite=_str_or_none(row[8]),
                nature_marchandises=_str_or_none(row[9]),
                observations=_str_or_none(row[10]),
            )
    finally:
        wb.close()

    return base


def resoudre_fournisseur(
    code: str,
    nom: str,
    base: dict[str, Fournisseur],
    delai_par_defaut: int = DELAI_PAR_DEFAUT_JOURS,
) -> tuple[Fournisseur, bool]:
    """Retourne le fournisseur depuis la base, ou en crée un avec le délai par défaut.

    Retourne (fournisseur, est_hors_base).
    """
    if code in base:
        return base[code], False
    return Fournisseur(code=code, nom=nom, delai_convenu_jours=delai_par_defaut), True
