"""Génération du fichier `Suivi Global DDP.xlsx` à partir des LigneSuivi.

Stratégie : on charge le template `Modèle Suivi Global.xlsx` (qui contient
déjà la mise en forme, l'en-tête, et les onglets attendus) et on remplit :

  - Onglet `Suivi Global DP`  : une ligne par LigneSuivi à partir de la ligne 10
  - Onglet `Base Frs Permanente` : référentiel fournisseurs utilisé

Les colonnes G/H/I/J du template contiennent des formules VLOOKUP basées sur
un Tableau structuré. Pour la phase 1, on inscrit directement les valeurs
calculées (plus robuste qu'essayer de propager les formules sur N lignes).

Le résultat est un fichier que le comptable peut ouvrir, ajuster, signer.
"""

from __future__ import annotations

import shutil
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from models import Fournisseur, LigneSuivi, StatutFacture  # noqa: E402


# ─── Configuration du rendu ────────────────────────────────────────────────

ONGLET_SUIVI = "Suivi Global DP"
ONGLET_BASE  = "Base Frs Permanente"

# Position de l'en-tête et début des données dans le template
LIGNE_DEBUT_DONNEES = 10

# Colonnes (1-based) — alignées sur le modèle de référence
COL_N_FACTURE       = 2   # B
COL_DATE_LIVRAISON  = 3   # C
COL_DATE_FACTURE    = 4   # D
COL_FOURNISSEUR     = 5   # E
COL_MONTANT_TTC     = 6   # F
COL_DELAI           = 7   # G
COL_ECHEANCE        = 8   # H
COL_RETARD          = 9   # I
COL_STATUT          = 10  # J
COL_DATE_PAIEMENT   = 11  # K
COL_OBSERVATIONS    = 12  # L

# Formats Excel
FORMAT_DATE         = "dd/mm/yyyy"
FORMAT_MONTANT      = "#,##0.00"
FORMAT_ENTIER       = "0"


# ─── Renderer principal ────────────────────────────────────────────────────

def generer_suivi_global(
    chemin_template: Path | str,
    chemin_sortie: Path | str,
    lignes: list[LigneSuivi],
    base_fournisseurs: dict[str, Fournisseur],
    client: str = "UEMA",
    exercice: int = 2026,
) -> Path:
    """Génère le fichier Suivi Global DDP à partir des lignes calculées.

    Le template est copié, les onglets `Suivi Global DP` et `Base Frs Permanente`
    sont remplis. Les autres onglets (Feuil1, SONDAGE, etc.) sont conservés tels quels.

    Retourne le chemin du fichier généré.
    """
    chemin_template = Path(chemin_template)
    chemin_sortie = Path(chemin_sortie)
    chemin_sortie.parent.mkdir(parents=True, exist_ok=True)

    # Copie du template (préserve mise en forme, onglets, tables)
    shutil.copy(chemin_template, chemin_sortie)

    wb = openpyxl.load_workbook(chemin_sortie)

    _remplir_suivi(wb[ONGLET_SUIVI], lignes, client, exercice)
    _remplir_base(wb[ONGLET_BASE], base_fournisseurs)

    wb.save(chemin_sortie)
    return chemin_sortie


# ─── Remplissage de l'onglet Suivi Global DP ───────────────────────────────

def _remplir_suivi(ws, lignes: list[LigneSuivi], client: str, exercice: int) -> None:
    """Remplit l'onglet Suivi Global DP."""
    # En-tête : Société + Exercice
    ws.cell(5, 4).value = client
    ws.cell(5, 12).value = exercice
    ws.cell(5, 12).number_format = "General"  # neutralise le format date du template

    # Récupérer le style de la ligne template (L10) pour le propager
    template_styles = _capturer_styles_ligne(ws, LIGNE_DEBUT_DONNEES)

    # Trier les lignes : par fournisseur puis date facture
    lignes_triees = sorted(lignes, key=lambda l: (l.fournisseur, l.date_facture, l.n_facture or ""))

    # Écrire chaque ligne
    for i, ligne in enumerate(lignes_triees):
        row = LIGNE_DEBUT_DONNEES + i
        _ecrire_ligne(ws, row, ligne)
        # Appliquer les styles capturés (sauf pour la première ligne déjà stylée)
        if i > 0:
            _appliquer_styles(ws, row, template_styles)
        # Flag visuel jaune sur les lignes "À vérifier"
        if ligne.observations and ligne.observations.startswith("⚠️"):
            _appliquer_flag_a_verifier(ws, row)


def _ecrire_ligne(ws, row: int, ligne: LigneSuivi) -> None:
    """Écrit une LigneSuivi à la ligne donnée."""
    ws.cell(row, COL_N_FACTURE).value      = ligne.n_facture
    ws.cell(row, COL_DATE_LIVRAISON).value = ligne.date_livraison
    ws.cell(row, COL_DATE_FACTURE).value   = ligne.date_facture
    ws.cell(row, COL_FOURNISSEUR).value    = ligne.fournisseur
    # Convention cabinet : pour un paiement partiel, on déclare uniquement
    # le reste à payer — le retard porte sur le reste, pas sur le total
    # facturé. Pour les autres statuts : montant TTC complet.
    if ligne.statut == StatutFacture.PARTIEL and ligne.montant_du is not None:
        montant_a_declarer = ligne.montant_du
    else:
        montant_a_declarer = ligne.montant_ttc
    ws.cell(row, COL_MONTANT_TTC).value    = float(montant_a_declarer)
    ws.cell(row, COL_DELAI).value          = ligne.delai_convenu_jours
    ws.cell(row, COL_ECHEANCE).value       = ligne.date_echeance
    # Convention cabinet : retard = échéance − paiement (négatif quand en retard).
    # Pour les OK RAS (payé à temps ou en avance), la cellule reste vide.
    if ligne.jours_retard is None:
        ws.cell(row, COL_RETARD).value = None
    elif ligne.jours_retard <= 0:
        # Payé à temps ou en avance → cellule vide (espace)
        ws.cell(row, COL_RETARD).value = " "
    else:
        # Retard : on affiche en négatif (convention cabinet)
        ws.cell(row, COL_RETARD).value = -ligne.jours_retard
    ws.cell(row, COL_STATUT).value         = ligne.statut.value
    ws.cell(row, COL_DATE_PAIEMENT).value  = ligne.date_paiement_effectif
    ws.cell(row, COL_OBSERVATIONS).value   = ligne.observations or ""

    # Formats Excel
    for col in (COL_DATE_LIVRAISON, COL_DATE_FACTURE, COL_ECHEANCE, COL_DATE_PAIEMENT):
        ws.cell(row, col).number_format = FORMAT_DATE
    ws.cell(row, COL_MONTANT_TTC).number_format = FORMAT_MONTANT
    ws.cell(row, COL_RETARD).number_format = FORMAT_ENTIER


# ─── Remplissage de la Base Frs ────────────────────────────────────────────

def _remplir_base(ws, base: dict[str, Fournisseur]) -> None:
    """Remplit l'onglet Base Frs Permanente.

    Structure du template :
      L1 : titre "BASE FOURNISSEURS PERMANENTE :"
      L3 : en-têtes "N° Fournisseur | Nom Fournisseur | Délai (jours) | Observations"
      L4+ : données
    """
    ligne_debut = 4
    fournisseurs_tries = sorted(base.values(), key=lambda f: f.nom)

    for i, f in enumerate(fournisseurs_tries):
        row = ligne_debut + i
        ws.cell(row, 1).value = f.code
        ws.cell(row, 2).value = f.nom
        ws.cell(row, 3).value = f.delai_convenu_jours
        ws.cell(row, 4).value = f.observations or ""


# ─── Gestion du style ──────────────────────────────────────────────────────

def _capturer_styles_ligne(ws, row: int) -> dict[int, dict]:
    """Capture les styles d'une ligne pour les répliquer."""
    styles = {}
    for col in range(COL_N_FACTURE, COL_OBSERVATIONS + 1):
        c = ws.cell(row, col)
        styles[col] = {
            "font": copy(c.font),
            "fill": copy(c.fill),
            "border": copy(c.border),
            "alignment": copy(c.alignment),
            "number_format": c.number_format,
        }
    return styles


def _appliquer_styles(ws, row: int, styles: dict[int, dict]) -> None:
    """Applique des styles préalablement capturés à une ligne."""
    for col, st in styles.items():
        c = ws.cell(row, col)
        c.font = st["font"]
        c.fill = st["fill"]
        c.border = st["border"]
        c.alignment = st["alignment"]
        # number_format est déjà fixé par _ecrire_ligne pour les colonnes spécifiques


# Fond jaune clair pour les lignes nécessitant une vérification manuelle
FILL_A_VERIFIER = PatternFill(
    start_color="FFF3CD",  # jaune pâle "warning"
    end_color="FFF3CD",
    fill_type="solid",
)


def _appliquer_flag_a_verifier(ws, row: int) -> None:
    """Met un fond jaune sur toute la ligne pour la signaler visuellement."""
    for col in range(COL_N_FACTURE, COL_OBSERVATIONS + 1):
        ws.cell(row, col).fill = FILL_A_VERIFIER
