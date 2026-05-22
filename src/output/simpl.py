"""Génération du formulaire DGI **Simpl délais de paiements** (CERFA ADC500B-23I).

Le Simpl est le formulaire officiel à déposer trimestriellement sur le portail
DGI Simpl-TVA. Il liste uniquement les **factures non payées dans les délais**
(FNP au 31 du trimestre + factures payées en retard durant le trimestre).

Structure (cf. référence UEMA 1T26) :

  L1-15  : En-tête identification (N° IF, Raison sociale, Période, Année,
           Chiffre d'affaires N-1, Taux BAM)
  L16    : Titre tableau "Etat des factures non payées dans les délais"
  L18    : Groupes de colonnes (Identification fournisseur | facture | …)
  L19    : En-têtes détaillées (31 colonnes)
  L20+   : Données — une ligne par facture

Stratégie phase 1 :
  - Le template est extrait du fichier de référence (en effaçant les données)
  - Les champs DGI manquants (N° IF, ICE, RC, adresse) sont laissés vides
    avec une observation pour le cabinet
"""

from __future__ import annotations

import math
import shutil
from copy import copy
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from models import Fournisseur, LigneSuivi, StatutFacture  # noqa: E402


# ─── Constantes du formulaire ──────────────────────────────────────────────

ONGLET = "Declaration délais paiement"
LIGNE_DEBUT_DONNEES = 20

# Cellules d'identification (en-tête)
CELL_N_IF       = (10, 2)   # L10 col B
CELL_PERIODE    = (10, 6)   # L10 col F
CELL_RAISON     = (11, 2)
CELL_ANNEE      = (11, 6)
CELL_ACTIVITE   = (12, 2)
CELL_CA_N1      = (12, 6)
CELL_TAUX_BAM   = (14, 2)

# Colonnes des données (1-based)
COL_FRS_IF          = 1
COL_FRS_ICE         = 2
COL_FRS_NOM         = 3
COL_FRS_RC          = 4
COL_FRS_ADRESSE     = 5
COL_FRS_VILLE       = 6
COL_FACT_N          = 7
COL_FACT_DATE_EMI   = 8
COL_FACT_NATURE     = 9
COL_LIV_DATE        = 10
COL_PAY_DATE_PREVUE = 14
COL_PAY_DATE_CONV   = 15
COL_MONTANT_TTC     = 18
COL_MONTANT_NON_PAYE = 19
COL_MOIS_RETARD_NP  = 20
COL_MONTANT_PAYE_HD = 21
COL_DATE_PAIEMENT_HD = 22
COL_MOIS_RETARD_PHD = 23
COL_MODE_PAIEMENT   = 29
COL_REF_PAIEMENT    = 30

FORMAT_DATE    = "dd/mm/yyyy"
FORMAT_MONTANT = "#,##0.00"
FORMAT_ENTIER  = "0"


# ─── Filtrage des lignes à déclarer ────────────────────────────────────────

# Statuts à inclure dans le Simpl (= « factures non payées dans les délais »)
STATUTS_DECLARABLES = {
    StatutFacture.NON_PAYE,
    StatutFacture.RETARD,
    StatutFacture.PARTIEL,
}


def filtrer_lignes_simpl(lignes: list[LigneSuivi]) -> list[LigneSuivi]:
    """Retient uniquement les lignes à reporter dans le Simpl."""
    return [l for l in lignes if l.statut in STATUTS_DECLARABLES]


# ─── Pipeline principal ────────────────────────────────────────────────────

def generer_simpl(
    chemin_template: Path | str,
    chemin_sortie: Path | str,
    lignes: list[LigneSuivi],
    base_fournisseurs: dict[str, Fournisseur],
    n_if_client: str,
    raison_sociale: str,
    periode_trimestre: int,                 # 1, 2, 3, 4
    annee: int,
    chiffre_affaires_n1: Optional[Decimal] = None,
    activite_code: Optional[int] = None,
    taux_bam: Decimal = Decimal("0.03"),
) -> Path:
    """Génère le fichier Simpl à partir des LigneSuivi et du référentiel fournisseurs."""
    chemin_template = Path(chemin_template)
    chemin_sortie = Path(chemin_sortie)
    chemin_sortie.parent.mkdir(parents=True, exist_ok=True)

    shutil.copy(chemin_template, chemin_sortie)
    wb = openpyxl.load_workbook(chemin_sortie)
    ws = wb[ONGLET]

    _remplir_entete(ws, n_if_client, raison_sociale, periode_trimestre, annee,
                    chiffre_affaires_n1, activite_code, taux_bam)

    lignes_a_declarer = filtrer_lignes_simpl(lignes)
    # Tri stable par fournisseur puis date facture
    lignes_a_declarer.sort(key=lambda l: (l.fournisseur, l.date_facture, l.n_facture or ""))

    _vider_zone_donnees(ws)
    template_styles = _capturer_styles_ligne(ws, LIGNE_DEBUT_DONNEES)

    for i, ligne in enumerate(lignes_a_declarer):
        row = LIGNE_DEBUT_DONNEES + i
        frs = base_fournisseurs.get(ligne.code_fournisseur or "")
        _ecrire_ligne(ws, row, ligne, frs)
        if i > 0:
            _appliquer_styles(ws, row, template_styles)

    wb.save(chemin_sortie)
    return chemin_sortie


# ─── Remplissage de l'en-tête ──────────────────────────────────────────────

def _remplir_entete(
    ws,
    n_if_client: str,
    raison_sociale: str,
    periode: int,
    annee: int,
    ca_n1: Optional[Decimal],
    activite: Optional[int],
    taux_bam: Decimal,
) -> None:
    ws.cell(*CELL_N_IF).value = n_if_client
    ws.cell(*CELL_PERIODE).value = periode
    ws.cell(*CELL_RAISON).value = raison_sociale
    ws.cell(*CELL_ANNEE).value = annee
    if activite is not None:
        ws.cell(*CELL_ACTIVITE).value = activite
    if ca_n1 is not None:
        ws.cell(*CELL_CA_N1).value = float(ca_n1)
    ws.cell(*CELL_TAUX_BAM).value = float(taux_bam)


# ─── Remplissage des lignes ────────────────────────────────────────────────

def _ecrire_ligne(ws, row: int, ligne: LigneSuivi, frs: Optional[Fournisseur]) -> None:
    """Écrit une ligne du Simpl à partir d'une LigneSuivi + des infos DGI du fournisseur."""

    # Identification fournisseur (cols 1-6) — depuis le référentiel
    if frs is not None:
        ws.cell(row, COL_FRS_IF).value      = frs.n_if or ""
        ws.cell(row, COL_FRS_ICE).value     = frs.n_ice or ""
        ws.cell(row, COL_FRS_NOM).value     = frs.nom
        ws.cell(row, COL_FRS_RC).value      = frs.n_rc or ""
        ws.cell(row, COL_FRS_ADRESSE).value = frs.adresse or ""
        ws.cell(row, COL_FRS_VILLE).value   = frs.ville_rc or ""
        ws.cell(row, COL_FACT_NATURE).value = frs.nature_marchandises or ""
    else:
        ws.cell(row, COL_FRS_NOM).value = ligne.fournisseur

    # Identification facture
    ws.cell(row, COL_FACT_N).value        = ligne.n_facture
    ws.cell(row, COL_FACT_DATE_EMI).value = ligne.date_facture

    # Livraison
    ws.cell(row, COL_LIV_DATE).value = ligne.date_livraison or ligne.date_facture

    # Échéance (date convenue)
    ws.cell(row, COL_PAY_DATE_CONV).value = ligne.date_echeance

    # Montant TTC
    ws.cell(row, COL_MONTANT_TTC).value = float(ligne.montant_ttc)

    # Différenciation FNP vs Retard
    if ligne.statut == StatutFacture.NON_PAYE:
        # Facture non encore payée
        ws.cell(row, COL_MONTANT_NON_PAYE).value = float(ligne.montant_ttc)
        # Mois de retard = aujourd'hui - échéance (approximation : on prend la fin de période)
        # En l'absence de date d'arrêté précise, on laisse vide — le cabinet remplit.
    elif ligne.statut in (StatutFacture.RETARD, StatutFacture.PARTIEL):
        # Facture payée hors délais
        ws.cell(row, COL_MONTANT_PAYE_HD).value = float(ligne.montant_ttc)
        if ligne.date_paiement_effectif:
            ws.cell(row, COL_DATE_PAIEMENT_HD).value = ligne.date_paiement_effectif
        if ligne.jours_retard is not None and ligne.jours_retard > 0:
            ws.cell(row, COL_MOIS_RETARD_PHD).value = math.ceil(ligne.jours_retard / 30)

    # Formats
    for col in (COL_FACT_DATE_EMI, COL_LIV_DATE, COL_PAY_DATE_CONV, COL_DATE_PAIEMENT_HD):
        ws.cell(row, col).number_format = FORMAT_DATE
    for col in (COL_MONTANT_TTC, COL_MONTANT_NON_PAYE, COL_MONTANT_PAYE_HD):
        ws.cell(row, col).number_format = FORMAT_MONTANT
    for col in (COL_MOIS_RETARD_NP, COL_MOIS_RETARD_PHD):
        ws.cell(row, col).number_format = FORMAT_ENTIER


# ─── Manipulations de la feuille ───────────────────────────────────────────

def _vider_zone_donnees(ws) -> None:
    """Efface les anciennes données (sous L19) sans toucher à la mise en forme."""
    for row in ws.iter_rows(min_row=LIGNE_DEBUT_DONNEES, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def _capturer_styles_ligne(ws, row: int) -> dict[int, dict]:
    styles = {}
    for col in range(1, 32):
        c = ws.cell(row, col)
        styles[col] = {
            "font": copy(c.font),
            "fill": copy(c.fill),
            "border": copy(c.border),
            "alignment": copy(c.alignment),
            "number_format": c.number_format,
        }
    return styles


def _appliquer_styles(ws, row: int, styles: dict[int, dict]) -> None:
    for col, st in styles.items():
        c = ws.cell(row, col)
        c.font = st["font"]
        c.fill = st["fill"]
        c.border = st["border"]
        c.alignment = st["alignment"]
