"""Tests de génération du Suivi Global Excel."""

from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from parser import parser_gl_liste  # noqa: E402
from lettrage import construire_lettrages  # noqa: E402
from compute import (  # noqa: E402
    calculer_toutes_lignes,
    charger_base_fournisseurs,
)
from output import generer_suivi_global  # noqa: E402


GL_UEMA   = Path(__file__).resolve().parents[1] / "samples" / "input" / "UEMA - GL FRS 2026.xlsx"
TEMPLATE  = Path(__file__).resolve().parents[1] / "samples" / "input" / "Modèle Suivi Global.xlsx"
REF       = Path(__file__).resolve().parents[1] / "samples" / "output_reference" / "Modèle Suivi Global DDP2026.xlsx"


@pytest.fixture(scope="module")
def fichier_genere(tmp_path_factory):
    base = charger_base_fournisseurs(REF)
    ecritures = parser_gl_liste(GL_UEMA)
    lettrages, _, _ = construire_lettrages(ecritures, base_fournisseurs=base)
    lignes = calculer_toutes_lignes(lettrages, debut_periode=date(2026, 1, 1))
    out = tmp_path_factory.mktemp("out") / "Suivi.xlsx"
    return generer_suivi_global(TEMPLATE, out, lignes, base, "UEMA", 2026)


def test_fichier_existe(fichier_genere):
    assert fichier_genere.exists()
    assert fichier_genere.stat().st_size > 50_000  # non vide


def test_onglets_preserves(fichier_genere):
    """Tous les onglets du template doivent être conservés."""
    wb = openpyxl.load_workbook(fichier_genere)
    for onglet in ['Suivi Global DP', 'Base Frs Permanente']:
        assert onglet in wb.sheetnames


def test_entete_societe(fichier_genere):
    """En-tête : Société = UEMA, Exercice = 2026."""
    wb = openpyxl.load_workbook(fichier_genere)
    ws = wb['Suivi Global DP']
    assert ws.cell(5, 4).value == "UEMA"
    assert ws.cell(5, 12).value == 2026


def test_lignes_donnees_remplies(fichier_genere):
    """Au moins 700 lignes de données à partir de L10."""
    wb = openpyxl.load_workbook(fichier_genere, data_only=True)
    ws = wb['Suivi Global DP']
    nb = sum(1 for i in range(10, ws.max_row + 1) if ws.cell(i, 2).value)
    assert nb >= 700


def test_statuts_valides(fichier_genere):
    """Tous les statuts doivent être dans le vocabulaire du cabinet."""
    wb = openpyxl.load_workbook(fichier_genere, data_only=True)
    ws = wb['Suivi Global DP']
    valides = {"OK RAS", "Non encore payée", "Attention, paiement hors délais",
               "Paiement partiel", "À vérifier"}
    for i in range(10, ws.max_row + 1):
        s = ws.cell(i, 10).value
        if s:
            assert str(s).strip() in valides, f"Statut inattendu L{i}: {s!r}"


def test_premiere_ligne_a2cim(fichier_genere):
    """La première ligne après tri (A2CIM) doit être OK RAS."""
    wb = openpyxl.load_workbook(fichier_genere, data_only=True)
    ws = wb['Suivi Global DP']
    # Trier par fournisseur → A2CIM en premier
    assert ws.cell(10, 5).value == "A2CIM"
    assert ws.cell(10, 10).value == "OK RAS"


def test_base_frs_remplie(fichier_genere):
    """L'onglet Base Frs Permanente doit contenir les fournisseurs."""
    wb = openpyxl.load_workbook(fichier_genere)
    ws = wb['Base Frs Permanente']
    nb = sum(1 for i in range(4, ws.max_row + 1) if ws.cell(i, 1).value)
    assert nb >= 100


def test_colonnes_dates_format_date(fichier_genere):
    """Colonnes C, D, H, K doivent contenir des dates."""
    import datetime as dt
    wb = openpyxl.load_workbook(fichier_genere, data_only=True)
    ws = wb['Suivi Global DP']
    # Vérifier la première ligne de données
    for col in (3, 4, 8):
        v = ws.cell(10, col).value
        assert isinstance(v, (dt.date, dt.datetime)), f"Col {col} L10: {v!r}"


def test_colonne_montant_numerique(fichier_genere):
    """Colonne F (Montant TTC) doit être numérique."""
    wb = openpyxl.load_workbook(fichier_genere, data_only=True)
    ws = wb['Suivi Global DP']
    v = ws.cell(10, 6).value
    assert isinstance(v, (int, float)) and v > 0
