"""Tests de génération du formulaire Simpl DGI."""

from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from parser import parser_gl_liste  # noqa: E402
from lettrage import construire_lettrages  # noqa: E402
from compute import calculer_toutes_lignes, charger_base_fournisseurs  # noqa: E402
from output import filtrer_lignes_simpl, generer_simpl  # noqa: E402
from models import LigneSuivi, StatutFacture  # noqa: E402


GL_UEMA = Path(__file__).resolve().parents[1] / "samples" / "input" / "UEMA - GL FRS 2026.xlsx"
REF_SUIVI = Path(__file__).resolve().parents[1] / "samples" / "output_reference" / "Modèle Suivi Global DDP2026.xlsx"
REF_SIMPL = Path(__file__).resolve().parents[1] / "samples" / "output_reference" / "Simpl délais de paiements UEMA -  TR 01-2026.xlsx"


# ─── Filtrage ──────────────────────────────────────────────────────────────

def test_filtrer_garde_fnp_et_retard():
    """Le filtre Simpl conserve FNP, Retard et Partiel — exclut OK_RAS."""
    from models import EcritureBrute, CodeJournal
    def _ligne(statut: StatutFacture) -> LigneSuivi:
        return LigneSuivi(
            n_facture="X", date_livraison=date(2026, 1, 1),
            date_facture=date(2026, 1, 1), fournisseur="TEST",
            montant_ttc=Decimal("100"), delai_convenu_jours=60,
            date_echeance=date(2026, 3, 2), jours_retard=None, statut=statut,
            date_paiement_effectif=None,
        )
    lignes = [_ligne(s) for s in StatutFacture]
    a_declarer = filtrer_lignes_simpl(lignes)
    statuts = {l.statut for l in a_declarer}
    assert statuts == {StatutFacture.NON_PAYE, StatutFacture.RETARD, StatutFacture.PARTIEL}
    assert StatutFacture.OK_RAS not in statuts


# ─── Génération sur UEMA ───────────────────────────────────────────────────

@pytest.fixture(scope="module")
def fichier_simpl(tmp_path_factory):
    base = charger_base_fournisseurs(REF_SUIVI)
    ecritures = parser_gl_liste(GL_UEMA)
    lettrages, _, _ = construire_lettrages(ecritures, base_fournisseurs=base)
    lignes = calculer_toutes_lignes(lettrages, debut_periode=date(2026, 1, 1))
    out = tmp_path_factory.mktemp("out") / "Simpl.xlsx"
    return generer_simpl(
        chemin_template=REF_SIMPL,
        chemin_sortie=out,
        lignes=lignes,
        base_fournisseurs=base,
        n_if_client="14367938",
        raison_sociale="STE UEMA INDUSTRY",
        periode_trimestre=1,
        annee=2026,
        chiffre_affaires_n1=Decimal("23369748.49"),
        activite_code=1,
    )


def test_fichier_simpl_existe(fichier_simpl):
    assert fichier_simpl.exists()
    assert fichier_simpl.stat().st_size > 30_000


def test_simpl_entete(fichier_simpl):
    wb = openpyxl.load_workbook(fichier_simpl, data_only=True)
    ws = wb["Declaration délais paiement"]
    assert ws.cell(10, 2).value == "14367938"
    assert ws.cell(10, 6).value == 1
    assert ws.cell(11, 2).value == "STE UEMA INDUSTRY"
    assert ws.cell(11, 6).value == 2026
    assert ws.cell(12, 6).value == pytest.approx(23369748.49)
    assert ws.cell(14, 2).value == pytest.approx(0.03)


def test_simpl_volume_dans_lordre(fichier_simpl):
    """Volume du Simpl dans une fourchette acceptable (cible 352)."""
    wb = openpyxl.load_workbook(fichier_simpl, data_only=True)
    ws = wb["Declaration délais paiement"]
    nb = sum(1 for i in range(20, ws.max_row + 1) if ws.cell(i, 3).value)
    assert 300 <= nb <= 700  # marge due à D-006 (dates AN)


def test_simpl_pas_de_ok_ras(fichier_simpl):
    """Aucune ligne du Simpl ne doit être un OK RAS (statut non-déclarable)."""
    wb = openpyxl.load_workbook(fichier_simpl, data_only=True)
    ws = wb["Declaration délais paiement"]
    # Si la facture est payée ET dans les délais, montant_payé_hd doit être vide
    # ou bien date_paiement_hd dans la période. Vérification approximative :
    # toute ligne doit avoir soit montant_non_paye (col 19), soit montant_paye_hd (col 21)
    nb_lignes = 0
    nb_correctes = 0
    for i in range(20, ws.max_row + 1):
        nom = ws.cell(i, 3).value
        if not nom:
            continue
        nb_lignes += 1
        non_paye = ws.cell(i, 19).value
        paye_hd  = ws.cell(i, 21).value
        if non_paye or paye_hd:
            nb_correctes += 1
    assert nb_correctes == nb_lignes  # chaque ligne a un montant à déclarer


def test_simpl_format_dates(fichier_simpl):
    """Les colonnes de date doivent contenir des dates."""
    import datetime as dt
    wb = openpyxl.load_workbook(fichier_simpl, data_only=True)
    ws = wb["Declaration délais paiement"]
    # première ligne de données
    for col in (8, 10, 15):  # date facture, livraison, échéance
        v = ws.cell(20, col).value
        if v is not None:
            assert isinstance(v, (dt.date, dt.datetime))


def test_simpl_montants_positifs(fichier_simpl):
    """Tous les montants TTC du Simpl doivent être strictement positifs."""
    wb = openpyxl.load_workbook(fichier_simpl, data_only=True)
    ws = wb["Declaration délais paiement"]
    for i in range(20, ws.max_row + 1):
        if ws.cell(i, 3).value:
            m = ws.cell(i, 18).value
            assert isinstance(m, (int, float)) and m > 0, f"L{i} montant={m!r}"


def test_simpl_pprime_dans_resultats(fichier_simpl):
    """PPRIME doit apparaître dans le Simpl (factures en retard connues)."""
    wb = openpyxl.load_workbook(fichier_simpl, data_only=True)
    ws = wb["Declaration délais paiement"]
    pprime_count = sum(
        1 for i in range(20, ws.max_row + 1)
        if ws.cell(i, 3).value and "PPRIME" in str(ws.cell(i, 3).value).upper()
    )
    assert pprime_count >= 5
