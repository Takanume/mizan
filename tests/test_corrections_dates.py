"""Tests du mécanisme de surcharge manuelle des dates (Phase 2 Sprint 2)."""

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from compute import (  # noqa: E402
    CorrectionDate,
    appliquer_correction,
    charger_corrections,
    construire_index,
    trouver_correction,
)
from compute.corrections import ONGLET_CORRECTIONS  # noqa: E402
from models import CodeJournal, EcritureBrute, Facture  # noqa: E402


# ─── Helpers ───────────────────────────────────────────────────────────────

def _ecrire_corrections_xlsx(path: Path, rows: list[tuple]) -> Path:
    """Crée un fichier de corrections au format attendu (7 cols : Frs, Nom, N°, Date, Livr, Obs, Délai)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ONGLET_CORRECTIONS
    ws["A1"] = "TITRE"
    ws["A2"] = "Notice"
    headers = ["N° Frs", "Nom", "N° Facture", "Date Facture", "Date Livr.", "Obs", "Délai"]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h).font = Font(bold=True)
    for i, row in enumerate(rows, start=4):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
    wb.save(path)
    return path


def _facture(code: str, nom: str, n_fact: str, d: date) -> Facture:
    ecr = EcritureBrute(
        ligne_source=0, code_fournisseur=code, nom_fournisseur=nom,
        date_ecriture=d, code_journal=CodeJournal.AN, n_piece="x",
        libelle=f"FN° {n_fact}", lettrage="A",
        debit=Decimal(0), credit=Decimal("100"), solde_progressif=None,
    )
    return Facture(
        code_fournisseur=code, nom_fournisseur=nom,
        n_piece="x", n_facture=n_fact,
        date_facture=d, date_livraison=None, montant_ttc=Decimal("100"),
        lettrage="A", is_report=True, source=ecr,
    )


# ─── Loader ────────────────────────────────────────────────────────────────

def test_charger_fichier_inexistant():
    """Si le fichier n'existe pas, retourne [] (pas d'erreur)."""
    assert charger_corrections(Path("/tmp/n_existe_pas.xlsx")) == []


def test_charger_corrections(tmp_path):
    path = _ecrire_corrections_xlsx(tmp_path / "c.xlsx", [
        ("FRS0001", "PPRIME", "25/023", datetime(2024, 11, 10), datetime(2025, 1, 31), "test"),
        ("FRS0002", "A2CIM",  "26FC0031", datetime(2026, 1, 23), None, None),
    ])
    cs = charger_corrections(path)
    assert len(cs) == 2
    assert cs[0].nom_fournisseur == "PPRIME"
    assert cs[0].n_facture == "25/023"
    assert cs[0].date_facture == date(2024, 11, 10)
    assert cs[0].date_livraison == date(2025, 1, 31)


def test_charger_ignore_lignes_incompletes(tmp_path):
    """Lignes sans n° facture ou sans date sont ignorées silencieusement."""
    path = _ecrire_corrections_xlsx(tmp_path / "c.xlsx", [
        ("FRS1", "PPRIME", "25/023", datetime(2024, 11, 10), None, None),
        ("FRS1", "PPRIME", None,     datetime(2024, 12, 1),  None, None),  # pas de n° → skip
        ("FRS1", "PPRIME", "25/024", None,                    None, None),  # pas de date → skip
        ("FRS1", None,     "25/025", datetime(2024, 12, 5),  None, None),  # pas de nom → skip
    ])
    cs = charger_corrections(path)
    assert len(cs) == 1
    assert cs[0].n_facture == "25/023"


# ─── Index ─────────────────────────────────────────────────────────────────

def test_index_par_code_et_nom():
    cs = [
        CorrectionDate("FRS0001", "PPRIME", "25/023", date(2024, 11, 10)),
        CorrectionDate(None,      "A2CIM",   "26FC0031", date(2026, 1, 23)),
    ]
    idx = construire_index(cs)
    assert ("FRS0001", "25/023") in idx
    assert ("PPRIME",  "25/023") in idx
    assert ("A2CIM",   "26FC0031") in idx


def test_trouver_correction_par_code():
    cs = [CorrectionDate("FRS0001", "PPRIME", "25/023", date(2024, 11, 10))]
    idx = construire_index(cs)
    f = _facture("FRS0001", "PPRIME", "25/023", date(2026, 1, 1))
    c = trouver_correction(f, idx)
    assert c is not None
    assert c.date_facture == date(2024, 11, 10)


def test_trouver_correction_par_nom_fallback():
    """Si le code ne match pas, on fallback sur le nom."""
    cs = [CorrectionDate("FRS_AUTRE", "PPRIME", "25/023", date(2024, 11, 10))]
    idx = construire_index(cs)
    f = _facture("FRS_DIFFERENT", "PPRIME", "25/023", date(2026, 1, 1))
    c = trouver_correction(f, idx)
    assert c is not None  # match par nom


def test_aucune_correction():
    idx = construire_index([])
    f = _facture("FRS0001", "PPRIME", "25/023", date(2026, 1, 1))
    assert trouver_correction(f, idx) is None


# ─── Application ───────────────────────────────────────────────────────────

def test_appliquer_correction():
    f = _facture("FRS0001", "PPRIME", "25/023", date(2026, 1, 1))
    c = CorrectionDate("FRS0001", "PPRIME", "25/023",
                       date(2024, 11, 10), date(2025, 1, 31))
    nouvelle = appliquer_correction(f, c)
    assert nouvelle.date_facture == date(2024, 11, 10)
    assert nouvelle.date_livraison == date(2025, 1, 31)
    # Le reste de la facture est préservé
    assert nouvelle.n_facture == f.n_facture
    assert nouvelle.code_fournisseur == f.code_fournisseur


def test_appliquer_correction_sans_date_livraison():
    """Si date_livraison non précisée, elle prend la date_facture."""
    f = _facture("FRS0001", "PPRIME", "25/023", date(2026, 1, 1))
    c = CorrectionDate("FRS0001", "PPRIME", "25/023", date(2024, 11, 10))
    nouvelle = appliquer_correction(f, c)
    assert nouvelle.date_livraison == date(2024, 11, 10)


# ─── Sprint 3 — Surcharge délai (D-003) ─────────────────────────────────────

def test_correction_porte_delai_surcharge():
    c = CorrectionDate(
        code_fournisseur="FRS0001", nom_fournisseur="PPRIME",
        n_facture="25/043", date_facture=date(2025, 2, 18),
        delai_jours=90,
    )
    f = _facture("FRS0001", "PPRIME", "25/043", date(2026, 1, 1))
    nouvelle = appliquer_correction(f, c)
    assert nouvelle.delai_surcharge == 90


def test_correction_sans_delai_garde_facture_intacte():
    """Si la correction n'a pas de délai, le délai surcharge reste None."""
    c = CorrectionDate(None, "PPRIME", "25/043", date(2025, 2, 18))
    f = _facture("FRS0001", "PPRIME", "25/043", date(2026, 1, 1))
    nouvelle = appliquer_correction(f, c)
    assert nouvelle.delai_surcharge is None


def test_loader_delai_optionnel(tmp_path):
    """Le loader Excel doit lire la colonne G (Délai spécifique) si présente."""
    path = _ecrire_corrections_xlsx(tmp_path / "c.xlsx", [
        ("FRS1", "PPRIME", "25/043", datetime(2025, 2, 18), None, "test", 90),
        ("FRS1", "PPRIME", "25/044", datetime(2025, 2, 19), None, None, None),
    ])
    cs = charger_corrections(path)
    assert len(cs) == 2
    by_n = {c.n_facture: c for c in cs}
    assert by_n["25/043"].delai_jours == 90
    assert by_n["25/044"].delai_jours is None


def test_loader_ignore_delai_invalide(tmp_path):
    """Délais aberrants (≤0 ou >365) ignorés silencieusement."""
    path = _ecrire_corrections_xlsx(tmp_path / "c.xlsx", [
        ("FRS1", "X", "1", datetime(2025, 1, 1), None, None, -5),    # invalide
        ("FRS1", "X", "2", datetime(2025, 1, 1), None, None, 999),   # invalide
        ("FRS1", "X", "3", datetime(2025, 1, 1), None, None, 75),    # valide
    ])
    cs = charger_corrections(path)
    by_n = {c.n_facture: c for c in cs}
    assert by_n["1"].delai_jours is None
    assert by_n["2"].delai_jours is None
    assert by_n["3"].delai_jours == 75


def test_calcul_echeance_avec_delai_surcharge():
    """L'échéance doit utiliser le délai surchargé de la facture si présent."""
    from compute import calculer_date_echeance
    from dataclasses import replace
    f = _facture("FRS1", "PPRIME", "25/043", date(2025, 1, 1))
    f_avec_surcharge = replace(f, delai_surcharge=90)
    # Sans surcharge : délai par défaut 60 jours
    ech_60 = calculer_date_echeance(f, 60)
    # Avec surcharge à 90 : on ignore le 60 passé en arg
    ech_90 = calculer_date_echeance(f_avec_surcharge, 60)
    # Convention J1 (D-009 révisée 2026-05-22) : échéance = date + délai − 1
    assert ech_60 == date(2025, 3, 1)
    assert ech_90 == date(2025, 3, 31)


# ─── Intégration end-to-end ────────────────────────────────────────────────

def test_correction_modifie_le_retard(tmp_path):
    """Confirme que la correction de date impacte bien le calcul du retard."""
    from datetime import date
    from parser import parser_gl_liste
    from lettrage import construire_lettrages
    from compute import calculer_toutes_lignes, charger_base_fournisseurs

    ROOT = Path(__file__).resolve().parents[1]
    base = charger_base_fournisseurs(
        ROOT / "samples" / "input" / "Référentiel DGI - Cabinet.xlsx"
    )
    ecritures = parser_gl_liste(ROOT / "samples" / "input" / "UEMA - GL FRS 2026.xlsx")

    # Sans corrections
    lettrages_a, _, _ = construire_lettrages(ecritures, base_fournisseurs=base)
    lignes_a = calculer_toutes_lignes(lettrages_a, debut_periode=date(2026, 1, 1))
    pprime_a = next((l for l in lignes_a
                       if l.code_fournisseur == "FRS0000235" and l.n_facture == "25/043"),
                      None)
    if pprime_a is None:
        pytest.skip("PPRIME 25/043 absent du GL — skip")

    # Avec corrections (vraie date 18/02/2025)
    path = _ecrire_corrections_xlsx(tmp_path / "c.xlsx", [
        ("FRS0000235", "PPRIME", "25/043", datetime(2025, 2, 18), None, None),
    ])
    cs = charger_corrections(path)
    idx = construire_index(cs)
    lettrages_b, _, _ = construire_lettrages(
        ecritures, base_fournisseurs=base, corrections_index=idx,
    )
    lignes_b = calculer_toutes_lignes(lettrages_b, debut_periode=date(2026, 1, 1))
    pprime_b = next((l for l in lignes_b
                       if l.code_fournisseur == "FRS0000235" and l.n_facture == "25/043"),
                      None)

    assert pprime_b is not None
    # La date a bien été corrigée
    assert pprime_b.date_facture == date(2025, 2, 18)
    # Et le retard a changé en conséquence
    assert pprime_a.jours_retard != pprime_b.jours_retard
